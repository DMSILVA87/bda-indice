"""
Pipeline para compilar dados de desenvolvimento públicos sobre Angola
a partir de instituições internacionais.

Saída: Angola_Development_Data.xlsx
Estrutura: 1 folha de Índice + 1 folha por fonte.
Período: 2000 - último disponível.
"""
from __future__ import annotations

import io
import json
import re
import time
import zipfile
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests

warnings.filterwarnings("ignore")

OUT_DIR = Path(r"e:\GitHub\Calisto Ebo")
OUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE = OUT_DIR / "_cache"
CACHE.mkdir(exist_ok=True)

ISO3 = "AGO"
ISO2 = "AO"
YEARS = list(range(2000, 2026))
YEAR_COLS = [str(y) for y in YEARS]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
}


def cache_get_insecure(url: str, name: str, timeout: int = 120, binary: bool = False) -> Optional[bytes | str]:
    """Fetch public statistical endpoints that have broken/intercepted CA chains."""
    p = CACHE / name
    if p.exists() and p.stat().st_size > 0:
        return p.read_bytes() if binary else p.read_text(encoding="utf-8", errors="replace")
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
        r.raise_for_status()
        data = r.content if binary else r.text
        if binary:
            p.write_bytes(data)
        else:
            p.write_text(data, encoding="utf-8", errors="replace")
        return data
    except Exception as e:
        print(f"   ! download falhou ({name}): {e}")
        return None


def cache_post_insecure(
    url: str,
    name: str,
    data: Dict[str, str],
    timeout: int = 120,
) -> Optional[str]:
    p = CACHE / name
    if p.exists() and p.stat().st_size > 0:
        return p.read_text(encoding="utf-8", errors="replace")
    try:
        r = requests.post(url, data=data, headers={**HEADERS, "Accept": "text/csv,*/*"},
                          timeout=timeout, verify=False)
        r.raise_for_status()
        p.write_text(r.text, encoding="utf-8", errors="replace")
        return r.text
    except Exception as e:
        print(f"   ! download falhou ({name}): {e}")
        return None


def cache_get(url: str, name: str, timeout: int = 120, binary: bool = False) -> Optional[bytes | str]:
    p = CACHE / name
    if p.exists() and p.stat().st_size > 0:
        return p.read_bytes() if binary else p.read_text(encoding="utf-8", errors="replace")
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        data = r.content if binary else r.text
        if binary:
            p.write_bytes(data)
        else:
            p.write_text(data, encoding="utf-8", errors="replace")
        return data
    except Exception as e:
        if "CERTIFICATE_VERIFY_FAILED" in str(e):
            try:
                r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
                r.raise_for_status()
                data = r.content if binary else r.text
                if binary:
                    p.write_bytes(data)
                else:
                    p.write_text(data, encoding="utf-8", errors="replace")
                return data
            except Exception as e2:
                print(f"   ! download falhou ({name}): {e2}")
                return None
        print(f"   ! download falhou ({name}): {e}")
        return None


def reshape_to_wide(
    df: pd.DataFrame,
    id_cols: List[str],
    year_col: str,
    value_col: str,
) -> pd.DataFrame:
    """Long -> wide with one column per year in YEARS."""
    df = df.copy()
    df[year_col] = pd.to_numeric(df[year_col], errors="coerce")
    df = df[df[year_col].between(min(YEARS), max(YEARS))]
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    wide = df.pivot_table(
        index=id_cols,
        columns=year_col,
        values=value_col,
        aggfunc="first",
    ).reset_index()
    wide.columns = [str(int(c)) if isinstance(c, (int, float)) and not pd.isna(c) else c for c in wide.columns]
    for y in YEAR_COLS:
        if y not in wide.columns:
            wide[y] = pd.NA
    return wide[id_cols + YEAR_COLS]


# ---------------------------------------------------------------------------
# 1) World Bank WDI — bulk
# ---------------------------------------------------------------------------
def fetch_wb_wdi() -> pd.DataFrame:
    print("[1/6] World Bank WDI — bulk download para AGO...")
    url = "https://api.worldbank.org/v2/en/country/AGO?downloadformat=csv"
    blob = cache_get(url, "wb_ago.zip", timeout=300, binary=True)
    if blob is None:
        return pd.DataFrame()
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile:
        print("   ! ficheiro WB inválido")
        return pd.DataFrame()

    data_name = next((n for n in z.namelist() if n.startswith("API_") and "Metadata" not in n), None)
    meta_name = next((n for n in z.namelist() if n.startswith("Metadata_Indicator")), None)
    if not data_name:
        return pd.DataFrame()

    with z.open(data_name) as f:
        raw = f.read().decode("utf-8", errors="replace")
    # skip the 4 header lines
    df = pd.read_csv(io.StringIO(raw), skiprows=4)

    year_cols = [c for c in df.columns if c.isdigit() and 2000 <= int(c) <= 2025]
    keep = ["Indicator Name", "Indicator Code"] + year_cols
    df = df[keep].copy()
    df = df.dropna(subset=year_cols, how="all")

    # add metadata: source organization (TOPIC was removed by WB in 2024)
    if meta_name:
        with z.open(meta_name) as f:
            mraw = f.read().decode("utf-8-sig", errors="replace")
        meta = pd.read_csv(io.StringIO(mraw))
        keep_meta = ["INDICATOR_CODE"]
        ren = {"INDICATOR_CODE": "Indicator Code"}
        if "SOURCE_ORGANIZATION" in meta.columns:
            keep_meta.append("SOURCE_ORGANIZATION")
            ren["SOURCE_ORGANIZATION"] = "Source Organization"
        if "SOURCE_NOTE" in meta.columns:
            keep_meta.append("SOURCE_NOTE")
            ren["SOURCE_NOTE"] = "Definition"
        meta = meta[keep_meta].rename(columns=ren)
        df = df.merge(meta, on="Indicator Code", how="left")
    if "Source Organization" not in df.columns:
        df["Source Organization"] = ""
    if "Definition" not in df.columns:
        df["Definition"] = ""

    df.insert(0, "Source", "World Bank — WDI")
    # ensure all 2000-2025 columns exist
    for y in YEAR_COLS:
        if y not in df.columns:
            df[y] = pd.NA
    df = df[["Source", "Indicator Code", "Indicator Name", "Source Organization", "Definition"] + YEAR_COLS]
    print(f"   ok — {len(df):,} indicadores")
    return df


# ---------------------------------------------------------------------------
# 2) IMF — WEO database (latest vintage)
# ---------------------------------------------------------------------------
def _parse_imf_sdmx_xml(xml_text: str, annual_only: bool = True) -> pd.DataFrame:
    """Parse IMF SDMX 2.1 structure-specific XML into long-form DataFrame.

    Output columns: Indicator Code, Unit, Indicator Key (full series key), year, value.
    For datasets with extra dimensions (BOP, IRFCL, ...) we encode all
    non-time series attributes into a synthetic Indicator Key for uniqueness.
    """
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return pd.DataFrame()
    rows = []
    drop_attrs = {"COUNTRY_UPDATE_DATE", "OVERLAP", "DECIMALS_DISPLAYED",
                  "DECIMALS", "OBS_STATUS", "TIME_FORMAT", "BASE_PER",
                  "UPDATE_DATE"}
    for series in root.iter():
        if not series.tag.endswith("Series"):
            continue
        attrs = dict(series.attrib)
        freq = attrs.get("FREQUENCY", attrs.get("FREQ", "A"))
        if annual_only and freq not in ("A", "ANNUAL", "ANN"):
            continue
        indicator = attrs.get("INDICATOR") or attrs.get("INDICATOR_CODE") or attrs.get("CONCEPT", "")
        unit = attrs.get("UNIT") or attrs.get("UNIT_MEASURE") or attrs.get("SCALE", "")
        # Build a deterministic key from attributes (excludes time/admin)
        key_parts = []
        for k, v in attrs.items():
            if k in drop_attrs or k in ("REF_AREA", "COUNTRY", "FREQ", "FREQUENCY"):
                continue
            if v:
                key_parts.append(f"{k}={v}")
        full_key = "|".join(key_parts) if key_parts else (indicator or "")
        for obs in series:
            if not obs.tag.endswith("Obs"):
                continue
            year = obs.attrib.get("TIME_PERIOD") or obs.attrib.get("TIME")
            value = obs.attrib.get("OBS_VALUE")
            if year is None or value is None:
                continue
            rows.append({
                "Indicator Code": indicator,
                "Indicator Key": full_key,
                "Unit": unit,
                "year": year,
                "value": value,
            })
    return pd.DataFrame(rows)


def _imf_codelist_names(agency: str = "IMF.RES", codelist: str = "CL_WEO_INDICATOR") -> Dict[str, str]:
    """Fetch indicator labels from IMF SDMX codelist."""
    url = f"https://api.imf.org/external/sdmx/2.1/codelist/{agency}/{codelist}"
    blob = cache_get(url, f"imf_cl_{agency}_{codelist}.xml", timeout=60)
    if not blob:
        return {}
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(blob)
    except ET.ParseError:
        return {}
    out = {}
    for code in root.iter():
        if not code.tag.endswith("Code"):
            continue
        cid = code.attrib.get("id")
        if not cid:
            continue
        for child in code:
            if child.tag.endswith("Name"):
                if (child.attrib.get("{http://www.w3.org/XML/1998/namespace}lang", "en") or "en") == "en":
                    out[cid] = (child.text or "").strip()
                    break
    return out


def fetch_imf_weo() -> pd.DataFrame:
    print("[2/6] IMF WEO — SDMX api.imf.org para AGO...")
    url = "https://api.imf.org/external/sdmx/2.1/data/IMF.RES,WEO/AGO?startPeriod=2000"
    blob = cache_get(url, "imf_weo_ago.xml", timeout=120)
    if not blob:
        return pd.DataFrame()
    df = _parse_imf_sdmx_xml(blob)
    if df.empty:
        return df
    # add labels
    names = _imf_codelist_names("IMF.RES", "CL_WEO_INDICATOR")
    df["Indicator Name"] = df["Indicator Code"].map(names).fillna(df["Indicator Code"])
    wide = reshape_to_wide(df, id_cols=["Indicator Code", "Indicator Name", "Unit"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "IMF — World Economic Outlook (WEO)")
    print(f"   ok — {len(wide)} séries WEO")
    return wide


def _fetch_imf_sta(dataset_id: str, version: str, label: str,
                   codelist: str = "CL_INDICATOR") -> pd.DataFrame:
    url = f"https://api.imf.org/external/sdmx/2.1/data/IMF.STA,{dataset_id},{version}/AGO?startPeriod=2000"
    blob = cache_get(url, f"imf_sta_{dataset_id}.xml", timeout=240)
    if not blob:
        return pd.DataFrame()
    df = _parse_imf_sdmx_xml(blob, annual_only=True)
    if df.empty:
        return df
    names = _imf_codelist_names("IMF.STA", codelist)
    df["Indicator Name"] = df["Indicator Code"].map(names).fillna(df["Indicator Code"])
    wide = reshape_to_wide(df, id_cols=["Indicator Code", "Indicator Name", "Indicator Key", "Unit"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", f"IMF — {label}")
    return wide


def fetch_imf_bop() -> pd.DataFrame:
    print("[2b/6] IMF STA — Balance of Payments (BOP) annual para AGO...")
    df = _fetch_imf_sta("BOP", "21.0.0", "Balance of Payments (BOP)", codelist="CL_BOP_INDICATOR")
    print(f"   ok — {len(df)} séries BOP")
    return df


def fetch_imf_cpi() -> pd.DataFrame:
    print("[2c/6] IMF STA — Consumer Price Index (CPI) annual para AGO...")
    df = _fetch_imf_sta("CPI", "5.0.0", "Consumer Price Index (CPI)", codelist="CL_CPI_INDICATOR")
    print(f"   ok — {len(df)} séries CPI")
    return df


def fetch_imf_irfcl() -> pd.DataFrame:
    print("[2d/6] IMF STA — International Reserves & FX Liquidity (IRFCL)...")
    df = _fetch_imf_sta("IRFCL", "11.0.0", "International Reserves & FX Liquidity (IRFCL)",
                         codelist="CL_IRFCL_INDICATOR")
    print(f"   ok — {len(df)} séries IRFCL")
    return df


def fetch_imf_anea() -> pd.DataFrame:
    print("[2e/6] IMF STA — National Economic Accounts (ANEA)...")
    df = _fetch_imf_sta("ANEA", "6.0.1", "National Economic Accounts, Annual (ANEA)",
                         codelist="CL_NEA_INDICATOR")
    print(f"   ok — {len(df)} séries ANEA")
    return df


def fetch_imf_ls() -> pd.DataFrame:
    print("[2f/6] IMF STA — Labor Statistics (LS)...")
    df = _fetch_imf_sta("LS", "9.0.0", "Labor Statistics (LS)", codelist="CL_LS_INDICATOR")
    print(f"   ok — {len(df)} séries LS")
    return df


# ---------------------------------------------------------------------------
# 3) UN HDI + UNCTAD
# ---------------------------------------------------------------------------
def fetch_undp_hdi() -> pd.DataFrame:
    print("[3a/6] UNDP HDR — Composite Indices CSV completo, filtrado AGO...")
    # Direct CSV from UNDP HDR data center (no API key required)
    url = "https://hdr.undp.org/sites/default/files/2023-24_HDR/HDR23-24_Composite_indices_complete_time_series.csv"
    blob = cache_get(url, "undp_hdr_full.csv", timeout=120)
    if not blob:
        return pd.DataFrame()
    try:
        df = pd.read_csv(io.StringIO(blob))
    except Exception as e:
        print(f"   ! parse failed: {e}")
        return pd.DataFrame()
    if "iso3" not in df.columns:
        return pd.DataFrame()
    df = df[df["iso3"] == "AGO"].copy()
    if df.empty:
        return pd.DataFrame()
    # Wide layout: columns like hdi_1990, hdi_1991, ... gii_1990, ...
    meta_cols = [c for c in ["iso3", "country", "hdicode", "region", "hdi_rank_2022"] if c in df.columns]
    data_cols = [c for c in df.columns if c not in meta_cols]
    rows = []
    for c in data_cols:
        # split indicator_year
        parts = c.rsplit("_", 1)
        if len(parts) != 2 or not parts[1].isdigit():
            continue
        ind, year = parts[0], int(parts[1])
        if not (2000 <= year <= 2025):
            continue
        val = df[c].iloc[0]
        if pd.isna(val):
            continue
        rows.append({"Indicator Code": ind, "year": year, "value": val})
    if not rows:
        return pd.DataFrame()
    rdf = pd.DataFrame(rows)
    # Human-readable labels
    labels = {
        "hdi": "Human Development Index (HDI)",
        "le": "Life expectancy at birth (years)",
        "eys": "Expected years of schooling (years)",
        "mys": "Mean years of schooling (years)",
        "gnipc": "GNI per capita (2017 PPP $)",
        "gdi": "Gender Development Index (GDI)",
        "hdi_f": "HDI — female",
        "hdi_m": "HDI — male",
        "le_f": "Life expectancy — female (years)",
        "le_m": "Life expectancy — male (years)",
        "eys_f": "Expected years of schooling — female",
        "eys_m": "Expected years of schooling — male",
        "mys_f": "Mean years of schooling — female",
        "mys_m": "Mean years of schooling — male",
        "gni_pc_f": "GNI per capita — female (2017 PPP $)",
        "gni_pc_m": "GNI per capita — male (2017 PPP $)",
        "gii": "Gender Inequality Index (GII)",
        "mmr": "Maternal mortality ratio (per 100,000 live births)",
        "abr": "Adolescent birth rate (per 1,000 women 15–19)",
        "se_f": "Population with at least some secondary education — female (%)",
        "se_m": "Population with at least some secondary education — male (%)",
        "pr_f": "Share of seats in parliament — female (%)",
        "pr_m": "Share of seats in parliament — male (%)",
        "lfpr_f": "Labour force participation rate — female (%)",
        "lfpr_m": "Labour force participation rate — male (%)",
        "ihdi": "Inequality-adjusted HDI (IHDI)",
        "coef_ineq": "Coefficient of human inequality",
        "loss": "Overall loss (%)",
        "ineq_le": "Inequality in life expectancy (%)",
        "ineq_edu": "Inequality in education (%)",
        "ineq_inc": "Inequality in income (%)",
        "phdi": "Planetary pressures-adjusted HDI (PHDI)",
        "diff_hdi_phdi": "Difference HDI vs PHDI (%)",
        "co2_prod": "CO2 emissions per capita (production, tonnes)",
        "mf": "Material footprint per capita (tonnes)",
        "rankdiff_hdi_phdi": "Rank difference HDI vs PHDI",
    }
    rdf["Indicator Name"] = rdf["Indicator Code"].map(labels).fillna(rdf["Indicator Code"])
    wide = reshape_to_wide(rdf, id_cols=["Indicator Code", "Indicator Name"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "UNDP — Human Development Reports (HDR 2023/24)")
    print(f"   ok — {len(wide)} séries UNDP HDR")
    return wide


def _clean_label(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _fetch_unctad_report(report: str) -> pd.DataFrame:
    meta_url = f"https://unctadstat-api.unctad.org/api/reportMetadata/{report}/en"
    meta_blob = cache_get_insecure(meta_url, f"unctad_meta_{report}.json", timeout=90)
    if not meta_blob:
        return pd.DataFrame()
    try:
        meta = json.loads(meta_blob)
    except json.JSONDecodeError:
        return pd.DataFrame()

    defaults = meta.get("defaults", {})
    axes = []
    for axis in ("rowAxe", "colAxe", "pageAxe"):
        axes.extend(defaults.get(axis, []) or [])

    fields: List[str] = []
    for axis in axes:
        field = axis.get("field")
        if field and field not in fields:
            fields.append(field)

    observations = defaults.get("observations", []) or []
    measure_info: Dict[str, Dict[str, object]] = {}
    for obs in observations:
        obs_code = str(obs.get("code", "")).strip()
        if not obs_code:
            continue
        field = f"M{obs_code}"
        selected_measure = next((m for m in obs.get("measures", []) if m.get("selected")), None)
        magnitude = float(selected_measure.get("magnitude") or 1) if selected_measure else 1.0
        label = selected_measure.get("label") if selected_measure else obs.get("label")
        measure_info[field] = {
            "label": label or obs.get("label") or field,
            "magnitude": magnitude,
        }

    if not fields or not measure_info:
        return pd.DataFrame()

    economy_field = next(
        (
            axis.get("field")
            for axis in axes
            if axis.get("label") == "ECONOMY"
            or axis.get("name") == "Economies"
            or axis.get("field") == "Economy"
        ),
        None,
    )
    if not economy_field:
        return pd.DataFrame()

    title = report
    info_url = f"https://unctadstat-api.unctad.org/api/datacenter/reports/{report}/info/en"
    info_blob = cache_get_insecure(info_url, f"unctad_info_{report}.json", timeout=60)
    if info_blob:
        try:
            title = json.loads(info_blob).get("Title") or report
        except json.JSONDecodeError:
            pass

    data_url = f"https://unctadstat-api.unctad.org/datamart-api/{report}/cur/Facts"
    form = {
        "$select": ",".join(fields + list(measure_info.keys())),
        "$filter": f"{economy_field}/Code eq '024'",
        "$format": "csv",
        "culture": "en",
    }
    data_blob = cache_post_insecure(data_url, f"unctad_data_{report}.csv", form, timeout=180)
    if not data_blob:
        return pd.DataFrame()
    try:
        df = pd.read_csv(io.StringIO(data_blob))
    except Exception:
        return pd.DataFrame()
    if df.empty or "Year" not in df.columns:
        return pd.DataFrame()

    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    df = df[df["Year"].between(min(YEARS), max(YEARS))]
    if df.empty:
        return pd.DataFrame()

    dim_fields = [f for f in fields if f not in (economy_field, "Year")]
    group_cols = [f"{f}_Code" if f"{f}_Code" in df.columns else f for f in dim_fields]
    grouped = df.groupby(group_cols, dropna=False) if group_cols else [((), df)]

    rows = []
    for keys, g0 in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        dim_parts = []
        code_parts = []
        for field, key in zip(dim_fields, keys):
            label_col = f"{field}_Label"
            code_col = f"{field}_Code"
            label = (
                _clean_label(g0[label_col].dropna().iloc[0])
                if label_col in g0.columns and g0[label_col].notna().any()
                else _clean_label(key)
            )
            code = (
                _clean_label(g0[code_col].dropna().iloc[0])
                if code_col in g0.columns and g0[code_col].notna().any()
                else _clean_label(key)
            )
            if code and code != "nan":
                code_parts.append(f"{field}={code}")
            if label and label != "nan":
                dim_parts.append(f"{field}: {label}")

        for measure_field, info in measure_info.items():
            value_col = f"{measure_field}_Value"
            if value_col not in g0.columns:
                continue
            g = g0.copy()
            g[value_col] = pd.to_numeric(g[value_col], errors="coerce")
            magnitude = float(info["magnitude"] or 1)
            if magnitude:
                g[value_col] = g[value_col] / magnitude
            g = g.dropna(subset=[value_col])
            if g.empty:
                continue

            label = _clean_label(info["label"])
            name = _clean_label(title)
            if dim_parts:
                name += " — " + " · ".join(dim_parts)
            name += f" — {label}"
            code = f"UNCTAD|{report}|{measure_field}"
            if code_parts:
                code += "|" + "|".join(code_parts)

            row = {
                "Source": "UNCTAD — UNCTADstat Data Centre",
                "Indicator Code": code,
                "Indicator Name": name,
                "Unit": label,
            }
            for _, rec in g.iterrows():
                year = int(rec["Year"]) if pd.notna(rec["Year"]) else None
                if year in YEARS:
                    row[str(year)] = rec[value_col]
            rows.append(row)

    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).drop_duplicates(subset=["Indicator Code"], keep="first")
    for y in YEAR_COLS:
        if y not in out.columns:
            out[y] = pd.NA
    return out[["Source", "Indicator Code", "Indicator Name", "Unit"] + YEAR_COLS]


def fetch_unctad() -> pd.DataFrame:
    print("[3b/6] UNCTAD — UNCTADstat Data Centre para AGO...")
    reports = [
        "US.TradeMerchTotal",
        "US.TradeMerchBalance",
        "US.TradeServCatTotal",
        "US.GoodsAndServicesBpm6",
        "US.GoodsAndServBalanceBpm6",
        "US.GDPTotal",
        "US.GDPComponent",
        "US.FdiFlowsStock",
        "US.Remittances",
        "US.CurrAccBalance",
        "US.PopTotal",
        "US.PCI",
        "US.MerchantFleet",
        "US.ContPortThroughput",
    ]
    frames = []
    for report in reports:
        df = _fetch_unctad_report(report)
        if not df.empty:
            frames.append(df)
    if not frames:
        print("   ! UNCTAD indisponível")
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    out = out.dropna(subset=YEAR_COLS, how="all")
    print(f"   ok — {len(out)} séries UNCTAD")
    return out


# ---------------------------------------------------------------------------
# 4) AfDB + OECD
# ---------------------------------------------------------------------------
def fetch_afdb() -> pd.DataFrame:
    print("[4a/6] AfDB Socio-Economic Database — AGO via DBnomics...")
    # The AfDB/OpenDataForAfrica site is Cloudflare-protected for scripts.
    # DBnomics exposes the same AfDB dataset through a stable JSON API.
    frames = []
    offset = 0
    limit = 1000
    num_found = None
    labels: Dict[str, Dict[str, str]] = {}
    while num_found is None or offset < num_found:
        url = (
            "https://api.db.nomics.world/v22/series/AFDB/bbkawjf/AGO"
            f"?observations=1&limit={limit}&offset={offset}"
        )
        blob = cache_get_insecure(url, f"afdb_dbnomics_ago_{offset}.json", timeout=180)
        if not blob:
            break
        try:
            payload = json.loads(blob)
        except json.JSONDecodeError:
            break

        dataset = payload.get("dataset", {})
        labels = dataset.get("dimensions_values_labels", labels) or labels
        series = payload.get("series", {})
        docs = series.get("docs", [])
        num_found = int(series.get("num_found", 0))
        if not docs:
            break

        rows = []
        for item in docs:
            dims = item.get("dimensions", {}) or {}
            indicator = dims.get("indicator", "")
            units = dims.get("units", "")
            scale = dims.get("scale", "")
            freq = dims.get("frequency", "")
            indicator_name = (
                labels.get("indicator", {}).get(indicator)
                or str(item.get("series_name", "")).replace(" - Angola", "")
                or indicator
            )
            unit_label = labels.get("units", {}).get(units, units)
            code = f"AFDB|{indicator}|UNIT={units}|SCALE={scale}|FREQ={freq}"
            row = {
                "Source": "AfDB — Socio Economic Database (via DBnomics)",
                "Indicator Code": code,
                "Indicator Name": indicator_name,
                "Unit": unit_label,
            }
            for period, value in zip(item.get("period", []), item.get("value", [])):
                if not str(period).isdigit():
                    continue
                year = int(period)
                if year not in YEARS:
                    continue
                numeric = pd.to_numeric(value, errors="coerce")
                if pd.notna(numeric):
                    row[str(year)] = numeric
            rows.append(row)

        if rows:
            frames.append(pd.DataFrame(rows))
        offset += limit

    if not frames:
        print("   ! AfDB indisponível")
        return pd.DataFrame()

    out = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["Indicator Code"], keep="first")
    for y in YEAR_COLS:
        if y not in out.columns:
            out[y] = pd.NA
    out = out[["Source", "Indicator Code", "Indicator Name", "Unit"] + YEAR_COLS]
    out = out.dropna(subset=YEAR_COLS, how="all")
    print(f"   ok — {len(out)} séries AfDB")
    return out


def fetch_oecd_oda() -> pd.DataFrame:
    print("[4b/6] OECD DAC — novo SDMX (sdmx.oecd.org) — ODA para AGO...")
    url = (
        "https://sdmx.oecd.org/public/rest/data/OECD.DCD.FSD,DSD_DAC2@DF_DAC2A,1.0/"
        ".AGO...?startPeriod=2000&dimensionAtObservation=AllDimensions"
    )
    blob = cache_get(url, "oecd_dac2a_ago.xml", timeout=240)
    if not blob:
        return pd.DataFrame()
    # The OECD response uses Generic SDMX. Stream parse with iterparse for memory.
    import xml.etree.ElementTree as ET
    rows = []
    try:
        # iterparse over bytes
        src = io.BytesIO(blob if isinstance(blob, bytes) else blob.encode("utf-8"))
        for event, elem in ET.iterparse(src, events=("end",)):
            tag = elem.tag.split("}", 1)[-1]
            if tag != "Obs":
                continue
            key_vals: Dict[str, str] = {}
            value = None
            for child in elem:
                ctag = child.tag.split("}", 1)[-1]
                if ctag == "ObsKey":
                    for v in child:
                        kid = v.attrib.get("id")
                        kv = v.attrib.get("value")
                        if kid and kv is not None:
                            key_vals[kid] = kv
                elif ctag == "ObsValue":
                    value = child.attrib.get("value")
                elif ctag == "Attributes":
                    for v in child:
                        kid = v.attrib.get("id")
                        kv = v.attrib.get("value")
                        if kid and kv is not None:
                            key_vals[kid] = kv
            year = key_vals.pop("TIME_PERIOD", None)
            if year is None or value is None:
                elem.clear()
                continue
            donor = key_vals.get("DONOR", "")
            measure = key_vals.get("MEASURE", "")
            flow = key_vals.get("FLOW_TYPE", "")
            price = key_vals.get("PRICE_BASE", "")
            unit = key_vals.get("UNIT_MEASURE", "USD")
            mult = key_vals.get("UNIT_MULT", "")
            code = f"DAC2A|D={donor}|M={measure}|F={flow}|P={price}"
            rows.append({
                "Indicator Code": code,
                "Indicator Name": f"Donor={donor} · Aid type={measure} · Flow={flow} · Price={price}",
                "Unit": f"{unit} (10^{mult})" if mult else unit,
                "year": year,
                "value": value,
            })
            elem.clear()
    except ET.ParseError as e:
        print(f"   ! XML parse error: {e}")
        return pd.DataFrame()
    if not rows:
        print("   ! OECD vazio após parse")
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # Many donor×measure×flow combos. Keep all but ensure uniqueness.
    df = df.drop_duplicates(subset=["Indicator Code", "year"], keep="first")
    wide = reshape_to_wide(df, id_cols=["Indicator Code", "Indicator Name", "Unit"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "OECD — DAC2A (ODA disbursements to Angola)")
    print(f"   ok — {len(wide)} séries OECD DAC")
    return wide


# ---------------------------------------------------------------------------
# 5) Others — ILO, WHO, FAO, Transparency
# ---------------------------------------------------------------------------
def fetch_ilostat() -> pd.DataFrame:
    print("[5a/6] ILOSTAT — rplumber API, bulk para AGO...")
    # rplumber gives us per-indicator data for a given country
    # Approach: fetch TOC -> for each annual indicator covering AGO, pull data
    # That's hundreds of calls. Instead, use ref_area filter on common indicators.
    toc_url = "https://rplumber.ilo.org/metadata/toc/indicator?lang=en&format=.csv"
    toc_blob = cache_get(toc_url, "ilo_toc.csv", timeout=120)
    if not toc_blob:
        return pd.DataFrame()
    try:
        toc = pd.read_csv(io.StringIO(toc_blob))
    except Exception:
        return pd.DataFrame()
    # Pick a curated subset (full TOC is too many calls). Core labor indicators:
    core_ids = [
        "EAP_DWAP_SEX_AGE_RT",       # Labour force participation rate
        "EMP_DWAP_SEX_AGE_RT",       # Employment-to-population ratio
        "UNE_DEAP_SEX_AGE_RT",       # Unemployment rate
        "EMP_TEMP_SEX_AGE_NB",       # Employment numbers
        "EAP_TEAP_SEX_AGE_NB",       # Labour force numbers
        "POP_2POP_SEX_AGE_NB",       # Working-age population
        "UNE_TUNE_SEX_AGE_NB",       # Unemployed persons
        "EMP_2EMP_SEX_AGE_ECO_NB",   # Employment by economic activity
        "EMP_TEMP_SEX_OCU_NB",       # Employment by occupation
        "EMP_TEMP_SEX_STE_NB",       # Employment by status
        "EAR_4MTH_SEX_CUR_NB",       # Mean monthly earnings
        "HOW_TEMP_SEX_NB",           # Hours of work
        "INJ_FATL_SEX_NB",           # Fatal occupational injuries
        "LAI_INDU_NOC_RT",           # Labour income share
        "EIP_TEIP_SEX_AGE_RT",       # NEET rate
        "GDP_205U_NOC_NB",           # GDP per worker
    ]
    rows = []
    for ind in core_ids:
        u = f"https://rplumber.ilo.org/data/indicator/?id={ind}&ref_area=AGO&type=both&format=.csv"
        bl = cache_get(u, f"ilo_{ind}.csv", timeout=90)
        if not bl or "time" not in bl[:500] and "TIME" not in bl[:500]:
            continue
        try:
            d = pd.read_csv(io.StringIO(bl))
        except Exception:
            continue
        if d.empty:
            continue
        ycol = "time" if "time" in d.columns else "TIME_PERIOD"
        vcol = "obs_value" if "obs_value" in d.columns else "OBS_VALUE"
        ind_label = d["indicator.label"].iloc[0] if "indicator.label" in d.columns else ind
        # Aggregate per year (mean of all breakdowns for headline)
        d = d[[ycol, vcol]].dropna()
        d[ycol] = pd.to_numeric(d[ycol], errors="coerce")
        d[vcol] = pd.to_numeric(d[vcol], errors="coerce")
        d = d.groupby(ycol, as_index=False)[vcol].mean()
        for _, r in d.iterrows():
            rows.append({
                "Indicator Code": ind,
                "Indicator Name": ind_label,
                "year": int(r[ycol]) if pd.notna(r[ycol]) else None,
                "value": r[vcol],
            })
    if not rows:
        print("   ! ILO bulk vazio")
        return pd.DataFrame()
    df = pd.DataFrame(rows).dropna(subset=["year"])
    wide = reshape_to_wide(df, id_cols=["Indicator Code", "Indicator Name"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "ILO — ILOSTAT (rplumber API)")
    print(f"   ok — {len(wide)} séries ILO")
    return wide


def fetch_who_gho() -> pd.DataFrame:
    print("[5b/6] WHO GHO — indicadores AGO...")
    # WHO GHO OData. Pull list of indicators first, then values for AGO.
    idx_url = "https://ghoapi.azureedge.net/api/Indicator"
    blob = cache_get(idx_url, "who_indicators.json", timeout=60)
    if not blob:
        return pd.DataFrame()
    try:
        idx = json.loads(blob).get("value", [])
    except json.JSONDecodeError:
        return pd.DataFrame()
    # WHO has thousands of indicators; pulling all would take too long.
    # Sample top 60 most-used categories. Use the dimension /api/Dimension.../DimensionValues
    # Simpler: pull a curated list of WHO indicators by code prefix.
    codes_of_interest = [
        "WHOSIS_000001",  # Life expectancy at birth
        "WHOSIS_000002",  # HALE
        "MDG_0000000001",  # Under-five mortality
        "MDG_0000000007",  # Infant mortality
        "MDG_0000000026",  # Maternal mortality
        "WHS4_543",  # Measles immunization
        "WHS4_100",  # DTP3 immunization
        "MALARIA_EST_INCIDENCE",
        "MALARIA_EST_DEATHS",
        "HIV_0000000026",  # HIV prevalence
        "MDG_0000000020",  # TB incidence
        "NCD_BMI_30A",  # Obesity adults
        "WSH_SANITATION_BASIC",  # Basic sanitation
        "WSH_WATER_BASIC",  # Basic water
        "SDGAIRBOD",  # Mortality from air pollution
        "GHED_CHEGDP_SHA2011",  # Current health exp % GDP
        "GHED_OOPSCHE_SHA2011",  # Out-of-pocket % current health exp
        "M_Est_smk_curr_std",  # Smoking adults
        "SA_0000001688",  # Alcohol consumption per capita
        "SDGSUICIDE",  # Suicide
    ]
    rows = []
    for code in codes_of_interest:
        u = f"https://ghoapi.azureedge.net/api/{code}?$filter=SpatialDim%20eq%20%27AGO%27"
        bl = cache_get(u, f"who_{code}.json", timeout=60)
        if not bl:
            continue
        try:
            vals = json.loads(bl).get("value", [])
        except json.JSONDecodeError:
            continue
        # find label
        label = next((it["IndicatorName"] for it in idx if it.get("IndicatorCode") == code), code)
        for v in vals:
            year = v.get("TimeDim")
            value = v.get("NumericValue")
            if year is None or value is None:
                continue
            rows.append({
                "Indicator Code": code,
                "Indicator Name": label,
                "year": int(year) if str(year).isdigit() else None,
                "value": value,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).dropna(subset=["year"])
    df["year"] = df["year"].astype(int)
    wide = reshape_to_wide(df, id_cols=["Indicator Code", "Indicator Name"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "WHO — Global Health Observatory")
    print(f"   ok — {len(wide)} séries WHO")
    return wide


def fetch_faostat() -> pd.DataFrame:
    print("[5c/6] FAOSTAT — produção agrícola AGO...")
    # FAOSTAT bulk: easier via macro-indicators country file.
    url = "https://bulks-faostat.fao.org/production/Macro-Statistics_Key_Indicators_E_All_Data_(Normalized).zip"
    blob = cache_get(url, "fao_macro.zip", timeout=300, binary=True)
    if not blob:
        return pd.DataFrame()
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile:
        return pd.DataFrame()
    csv_name = next((n for n in z.namelist() if n.endswith(".csv") and "Normalized" in n), None)
    if not csv_name:
        return pd.DataFrame()
    with z.open(csv_name) as f:
        raw = f.read().decode("latin-1", errors="replace")
    df = pd.read_csv(io.StringIO(raw), low_memory=False)
    if "Area" not in df.columns:
        return pd.DataFrame()
    df = df[df["Area"] == "Angola"].copy()
    if df.empty:
        return pd.DataFrame()
    df = df.rename(columns={"Year": "year", "Value": "value",
                            "Item": "Indicator Name", "Element": "Element"})
    df["Indicator Name"] = df["Indicator Name"].astype(str) + " — " + df["Element"].astype(str)
    df["Indicator Code"] = "FAO:" + df.get("Item Code", "").astype(str) + ":" + df.get("Element Code", "").astype(str)
    wide = reshape_to_wide(df[["Indicator Code", "Indicator Name", "year", "value"]],
                           id_cols=["Indicator Code", "Indicator Name"],
                           year_col="year", value_col="value")
    wide.insert(0, "Source", "FAO — FAOSTAT (Macro Indicators)")
    print(f"   ok — {len(wide)} séries FAO")
    return wide


def fetch_ti_cpi() -> pd.DataFrame:
    """Transparency International — Corruption Perceptions Index (manual data)."""
    print("[5d/6] Transparency Int. — CPI Angola (compilação manual)...")
    # TI CPI scores 2012-2024 for Angola (public, manually compiled from annual reports)
    cpi = {
        2012: 22, 2013: 23, 2014: 19, 2015: 15, 2016: 18, 2017: 19,
        2018: 19, 2019: 26, 2020: 27, 2021: 29, 2022: 33, 2023: 33, 2024: 32,
    }
    rank = {
        2012: 157, 2013: 153, 2014: 161, 2015: 163, 2016: 164, 2017: 167,
        2018: 165, 2019: 146, 2020: 142, 2021: 136, 2022: 116, 2023: 121, 2024: 121,
    }
    rows = [
        {"Indicator Code": "TI_CPI_SCORE", "Indicator Name": "Corruption Perceptions Index — score (0=highly corrupt, 100=clean)", **{str(y): v for y, v in cpi.items()}},
        {"Indicator Code": "TI_CPI_RANK", "Indicator Name": "Corruption Perceptions Index — global rank", **{str(y): v for y, v in rank.items()}},
    ]
    df = pd.DataFrame(rows)
    for y in YEAR_COLS:
        if y not in df.columns:
            df[y] = pd.NA
    df.insert(0, "Source", "Transparency International — CPI")
    return df[["Source", "Indicator Code", "Indicator Name"] + YEAR_COLS]


# ---------------------------------------------------------------------------
# Build Index sheet
# ---------------------------------------------------------------------------
def build_index(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for sheet, df in sheets.items():
        if df.empty:
            rows.append({
                "Folha": sheet,
                "Fonte": sheet,
                "Nº indicadores": 0,
                "Nº observações (não-nulas)": 0,
                "Período coberto": "—",
                "Notas": "Sem dados disponíveis nesta extração",
            })
            continue
        year_cols = [c for c in df.columns if c in YEAR_COLS]
        non_null = df[year_cols].notna().sum().sum()
        years_with_data = [int(c) for c in year_cols if df[c].notna().any()]
        period = f"{min(years_with_data)}–{max(years_with_data)}" if years_with_data else "—"
        source_name = df["Source"].iloc[0] if "Source" in df.columns and not df.empty else sheet
        rows.append({
            "Folha": sheet,
            "Fonte": source_name,
            "Nº indicadores": len(df),
            "Nº observações (não-nulas)": int(non_null),
            "Período coberto": period,
            "Notas": "",
        })
    idx = pd.DataFrame(rows)
    return idx


def main():
    sheets: Dict[str, pd.DataFrame] = {}

    sheets["WB_WDI"] = fetch_wb_wdi()
    sheets["IMF_WEO"] = fetch_imf_weo()
    sheets["IMF_BOP"] = fetch_imf_bop()
    sheets["IMF_CPI"] = fetch_imf_cpi()
    sheets["IMF_IRFCL"] = fetch_imf_irfcl()
    sheets["IMF_ANEA"] = fetch_imf_anea()
    sheets["IMF_LS"] = fetch_imf_ls()
    sheets["UN_UNDP_HDI"] = fetch_undp_hdi()
    sheets["UN_UNCTAD"] = fetch_unctad()
    sheets["AfDB"] = fetch_afdb()
    sheets["OECD_DAC"] = fetch_oecd_oda()
    sheets["ILO_ILOSTAT"] = fetch_ilostat()
    sheets["WHO_GHO"] = fetch_who_gho()
    sheets["FAO_FAOSTAT"] = fetch_faostat()
    sheets["TI_CPI"] = fetch_ti_cpi()

    out_path = OUT_DIR / "Angola_Development_Data.xlsx"
    try:
        with out_path.open("a+b"):
            pass
    except PermissionError:
        out_path = OUT_DIR / "Angola_Development_Data_completed.xlsx"
        print(f"\n   ! ficheiro principal bloqueado; a escrever cópia em {out_path}")
    print(f"\n[6/6] A escrever {out_path}...")
    idx = build_index(sheets)

    # Methodology sheet
    from datetime import date
    methodology = pd.DataFrame([
        {"Campo": "Título", "Valor": "Dados de desenvolvimento sobre Angola — compilação de instituições internacionais"},
        {"Campo": "País", "Valor": "Angola (ISO3: AGO)"},
        {"Campo": "Período coberto", "Valor": "2000 – mais recente disponível por fonte"},
        {"Campo": "Data da extração", "Valor": date.today().isoformat()},
        {"Campo": "Fontes incluídas (com dados)", "Valor":
            "World Bank WDI · IMF (WEO, BOP, CPI, IRFCL, ANEA, LS) · UNDP HDR · UNCTADstat · AfDB Socio Economic Database · OECD DAC · ILO ILOSTAT · WHO GHO · FAO FAOSTAT · Transparency International"},
        {"Campo": "Fontes tentadas sem retorno", "Valor":
            "AfDB/OpenDataForAfrica direto continua bloqueado por Cloudflare para acesso programático; "
            "foi usado o endpoint estruturado DBnomics para a base AfDB Socio Economic Database."},
        {"Campo": "Estrutura", "Valor":
            "1 folha 'Índice' (resumo por fonte) · 1 folha 'Metodologia' (este painel) · 1 folha por fonte. "
            "Em cada folha de dados: linhas = indicadores; colunas = anos 2000–2025."},
        {"Campo": "Endpoints utilizados", "Valor":
            "WB: api.worldbank.org bulk CSV · IMF: api.imf.org SDMX 2.1 · UNDP: hdr.undp.org composite indices CSV · "
            "UNCTAD: unctadstat-api.unctad.org Data Centre/OData · AfDB: api.db.nomics.world AFDB/bbkawjf · "
            "OECD: sdmx.oecd.org SDMX 2.1 generic format · ILO: rplumber.ilo.org · "
            "WHO: ghoapi.azureedge.net OData · FAO: bulks-faostat.fao.org Macro Indicators · "
            "TI: compilação manual de relatórios CPI anuais"},
        {"Campo": "Notas",
         "Valor":
            "• Frequência: anual (séries trimestrais/mensais agregadas pelo provedor). "
            "• Unidades preservadas como reportadas pela fonte (coluna 'Unit' onde disponível). "
            "• IMF IRFCL só cobre 2020+ porque a metodologia atual começou nessa altura. "
            "• OECD DAC2A: cada série é identificada por combinação Donor × Aid type × Flow × Price base."},
    ])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        idx.to_excel(xw, sheet_name="Índice", index=False)
        methodology.to_excel(xw, sheet_name="Metodologia", index=False)
        for name, df in sheets.items():
            sheet_name = name[:31]
            if df.empty:
                stub = pd.DataFrame({
                    "Aviso": [
                        f"Sem dados extraídos para '{name}' nesta execução.",
                        "Ver folha 'Metodologia' para detalhes sobre fontes indisponíveis.",
                    ]
                })
                stub.to_excel(xw, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(xw, sheet_name=sheet_name, index=False)

    # Light formatting: freeze top row + autofit-ish column widths
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"  # freeze first row
        # Widen first column to 35, year columns to 11, others to 25
        for i, col in enumerate(ws.columns, start=1):
            letter = col[0].column_letter
            header = (ws.cell(row=1, column=i).value or "")
            if header in YEAR_COLS:
                ws.column_dimensions[letter].width = 11
            elif i == 1:
                ws.column_dimensions[letter].width = 30
            else:
                # Auto-ish: use header length but cap
                ws.column_dimensions[letter].width = min(max(len(str(header)) + 2, 18), 45)
    wb.save(out_path)

    print(f"\nFICHEIRO ESCRITO: {out_path}")
    print(f"Tamanho: {out_path.stat().st_size/1024/1024:.1f} MB")
    print("\nResumo:")
    print(idx.to_string(index=False))


if __name__ == "__main__":
    main()
