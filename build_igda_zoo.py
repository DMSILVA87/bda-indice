from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import math
import re

import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(".")
BASE_FILE = ROOT / "IGDA_v3_Media_Geometrica_compilado_indicadores.xlsx"
V4_FILE = ROOT / "IGDA_v4_Metodologia_Linkada_Melhores_Indicadores.xlsx"
DATA_FILE = ROOT / "Calisto Ebo" / "Angola_Development_Data_completed.xlsx"
OUT_XLS = ROOT / "IGDA_BDA_zoo_indicadores_Angola_2026-06-02.xls"
OUT_XLSX = ROOT / "IGDA_BDA_zoo_indicadores_Angola_2026-06-02.xlsx"

YEARS = [str(y) for y in range(2000, 2026)]
BASE_YEARS = [str(y) for y in range(2015, 2026)]

DIM_BY_NUM = {
    "1": "1. Governança e Estado de Direito",
    "2": "2. Estabilidade Macroeconómica",
    "3": "3. Capital Humano",
    "4": "4. Inclusão Social e Protecção",
    "5": "5. Infraestruturas e Serviços",
    "6": "6. Mercado de Trabalho",
    "7": "7. Segurança Alimentar e Saúde",
    "8": "8. Diversificação Produtiva e Setor Privado",
    "9": "9. Ambiente, Clima e Resiliência",
    "10": "10. Demografia, Território e Urbanização",
    "11": "11. Transformação Digital e Inovação",
}

DIM_SHEETS = {
    DIM_BY_NUM["1"]: "01_Governanca",
    DIM_BY_NUM["2"]: "02_Macro",
    DIM_BY_NUM["3"]: "03_Capital_Humano",
    DIM_BY_NUM["4"]: "04_Inclusao",
    DIM_BY_NUM["5"]: "05_Infra_Serv",
    DIM_BY_NUM["6"]: "06_Mercado_Trabalho",
    DIM_BY_NUM["7"]: "07_Seg_Alim_Saude",
    DIM_BY_NUM["8"]: "08_Diversificacao",
    DIM_BY_NUM["9"]: "09_Ambiente_Clima",
    DIM_BY_NUM["10"]: "10_Demografia_Terr",
    DIM_BY_NUM["11"]: "11_Digital_Inov",
}

DIM_PREFIX = {
    DIM_BY_NUM["1"]: "GOV",
    DIM_BY_NUM["2"]: "MAC",
    DIM_BY_NUM["3"]: "HUM",
    DIM_BY_NUM["4"]: "INC",
    DIM_BY_NUM["5"]: "INF",
    DIM_BY_NUM["6"]: "LAB",
    DIM_BY_NUM["7"]: "SAU",
    DIM_BY_NUM["8"]: "DIV",
    DIM_BY_NUM["9"]: "AMB",
    DIM_BY_NUM["10"]: "DEM",
    DIM_BY_NUM["11"]: "DIG",
}

HEADERS = [
    "ID",
    "Dimensão",
    "Subtema",
    "Indicador",
    "Tipo",
    "Prioridade",
    "Sentido",
    "Unidade",
    "Mínimo",
    "Máximo",
    "Meta 2027",
    "Peso sugerido",
    "Código/API",
    "Fonte",
    "Fonte_URL",
    "Estado da disponibilidade",
    "Cobertura 2000-2025",
    "Cobertura 2015-2025",
    "Último ano obs.",
    "Origem no ficheiro",
    "Justificação / nota",
] + YEARS


EXTRAS = r"""
1|||Instituições públicas|||CPIA - direitos de propriedade e governação baseada em regras|||IQ.CPA.PROP.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Complementa WGI em qualidade institucional para países elegíveis IDA.
1|||Instituições públicas|||CPIA - transparência, responsabilização e corrupção no setor público|||IQ.CPA.TRAN.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Complementa CPI/WGI com avaliação institucional operacional.
1|||Instituições públicas|||CPIA - gestão do setor público e instituições|||IQ.CPA.PUBS.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Cluster institucional útil para robustez do capítulo de governação.
1|||Instituições públicas|||CPIA - qualidade da administração pública|||IQ.CPA.PADM.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Capta capacidade administrativa do Estado.
1|||Finanças públicas|||CPIA - qualidade da gestão orçamental e financeira|||IQ.CPA.FINQ.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Complementa transparência e execução orçamental.
1|||Mobilização de receita|||CPIA - eficiência da mobilização de receita|||IQ.CPA.REVN.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Pesquisa externa/compilação local|||Média|||Relevante para sustentabilidade do Estado.
1|||Corrupção empresarial|||Incidência de subornos em empresas|||IC.FRM.BRIB.ZS|||Banco Mundial/Enterprise Surveys|||%|||-|||0|||100|||10|||Pesquisa externa/compilação local|||Alta|||Capta corrupção sentida pelas empresas, complementar ao CPI/WGI.
1|||Corrupção empresarial|||Pagamentos informais a funcionários públicos|||IC.FRM.CORR.ZS|||Banco Mundial/Enterprise Surveys|||%|||-|||0|||100|||10|||Pesquisa externa/compilação local|||Alta|||Indicador empresarial direto de informalidade/corrupção administrativa.
1|||Segurança e paz|||Homicídios intencionais|||VC.IHR.PSRC.P5|||Banco Mundial/WDI|||por 100.000 hab.|||-|||0|||50|||5|||Pesquisa externa/compilação local|||Média|||Aproxima segurança pública; deve ser lido com qualidade do registo criminal.
1|||Segurança e paz|||Mortes relacionadas com conflito armado|||VC.BTL.DETH|||Banco Mundial/WDI|||número|||-|||0|||5000|||0|||Pesquisa externa/compilação local|||Média|||Inclui risco de estabilidade e segurança humana.
1|||Governança africana|||IIAG - Accountability and Transparency|||-|||Mo Ibrahim Foundation|||0-100|||+|||0|||100|||60|||Benchmark externo/manual|||Média|||Benchmark africano para validação regional do capítulo de governação.
1|||Governança africana|||IIAG - Security and Rule of Law|||-|||Mo Ibrahim Foundation|||0-100|||+|||0|||100|||60|||Benchmark externo/manual|||Média|||Benchmark africano alinhado com estado de direito e segurança.
2|||Finanças públicas|||Dívida bruta do governo geral - FMI WEO|||GGXWDG_NGDP|||FMI/WEO|||% PIB|||-|||0|||150|||60|||Pesquisa externa/compilação local|||Alta|||Série FMI alternativa/validação da dívida pública.
2|||Finanças públicas|||Receita do governo geral - FMI WEO|||GGR_NGDP|||FMI/WEO|||% PIB|||+|||0|||50|||25|||Pesquisa externa/compilação local|||Média|||Complementa receita fiscal WDI.
2|||Finanças públicas|||Despesa do governo geral - FMI WEO|||GGX_NGDP|||FMI/WEO|||% PIB|||+/-|||0|||60|||-|||Pesquisa externa/compilação local|||Baixa|||Indicador de contexto; interpretar com composição e eficiência da despesa.
2|||Finanças públicas|||Saldo líquido / necessidade de financiamento do governo geral|||GGXCNL_NGDP|||FMI/WEO|||% PIB|||+|||-20|||20|||0|||Pesquisa externa/compilação local|||Alta|||Alternativa FMI para saldo orçamental.
2|||Finanças públicas|||Saldo primário do governo geral|||GGXONLB_NGDP|||FMI/WEO|||% PIB|||+|||-20|||20|||0|||Pesquisa externa/compilação local|||Média|||Ajuda a separar juros da posição fiscal primária.
2|||Poupança e investimento|||Poupança nacional bruta|||NGSD_NGDP|||FMI/WEO|||% PIB|||+|||0|||60|||25|||Pesquisa externa/compilação local|||Média|||Sinal de capacidade interna de financiamento.
2|||Poupança e investimento|||Formação bruta de capital|||NID_NGDP|||FMI/WEO|||% PIB|||+|||0|||60|||25|||Pesquisa externa/compilação local|||Média|||Complementa investimento e crescimento potencial.
2|||Condições financeiras|||Taxa de juro real|||FR.INR.RINR|||Banco Mundial/WDI|||%|||+/-|||-20|||50|||-|||Pesquisa externa/compilação local|||Baixa|||Indicador de contexto financeiro; extremos positivos ou negativos podem ser problemáticos.
2|||Condições financeiras|||Taxa de juro ativa|||FR.INR.LEND|||Banco Mundial/WDI|||%|||-|||0|||80|||15|||Pesquisa externa/compilação local|||Média|||Proxy de custo de crédito.
2|||Condições financeiras|||Moeda ampla|||FM.LBL.BMNY.GD.ZS|||Banco Mundial/WDI|||% PIB|||+|||0|||100|||35|||Pesquisa externa/compilação local|||Média|||Indicador de profundidade monetária.
2|||Condições financeiras|||Taxa básica do BNA|||-|||BNA|||%|||-|||0|||40|||10|||Ficheiro nacional/manual|||Alta|||Disponível nos ficheiros BNA de taxas de juro do banco central; requer harmonização mensal-anual.
2|||Inflação|||Inflação homóloga nacional - BNA/INE|||-|||BNA/INE|||%|||-|||0|||50|||10|||Ficheiro nacional/manual|||Alta|||Há dataset BNA local em M1-MAcro/BNA/Preços e Contas Nacionais/Evolução Mensal da Taxa de Inflação.
3|||Escolaridade|||Anos esperados de escolaridade - PNUD|||eys|||PNUD/HDR|||anos|||+|||0|||18|||12|||Pesquisa externa/compilação local|||Alta|||Subcomponente clássico do HDI.
3|||Escolaridade|||Anos médios de escolaridade - PNUD|||mys|||PNUD/HDR|||anos|||+|||0|||15|||8|||Pesquisa externa/compilação local|||Alta|||Subcomponente clássico do HDI.
3|||Desigualdade humana|||Índice de Desenvolvimento Humano ajustado à desigualdade|||ihdi|||PNUD/HDR|||0-1|||+|||0|||1|||0.55|||Pesquisa externa/compilação local|||Média|||Aproxima perdas por desigualdade no desenvolvimento humano.
3|||Desigualdade humana|||Desigualdade na educação|||ineq_edu|||PNUD/HDR|||%|||-|||0|||100|||25|||Pesquisa externa/compilação local|||Média|||Complementa acesso e qualidade educativa.
3|||Literacia|||Alfabetização jovem|||SE.ADT.1524.LT.ZS|||Banco Mundial/UNESCO|||%|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Alta|||Capta aquisição recente de competências básicas.
3|||Literacia|||Alfabetização adulta feminina|||SE.ADT.LITR.FE.ZS|||Banco Mundial/UNESCO|||%|||+|||0|||100|||75|||Pesquisa externa/compilação local|||Alta|||Cruza capital humano e igualdade de género.
3|||Acesso escolar|||Crianças fora da escola primária|||SE.PRM.UNER.ZS|||Banco Mundial/UNESCO|||%|||-|||0|||100|||5|||Pesquisa externa/compilação local|||Alta|||Indicador de exclusão escolar.
3|||Acesso escolar|||Adolescentes fora do ensino secundário inferior|||SE.SEC.UNER.LO.ZS|||Banco Mundial/UNESCO|||%|||-|||0|||100|||10|||Pesquisa externa/compilação local|||Média|||Capta transição pós-primária.
3|||Conclusão escolar|||Conclusão do ensino secundário inferior|||SE.SEC.CMPT.LO.ZS|||Banco Mundial/UNESCO|||%|||+|||0|||100|||75|||Pesquisa externa/compilação local|||Média|||Indicador de stock de competências básicas alargadas.
3|||Qualidade docente|||Professores qualificados no ensino primário|||SE.PRM.TCAQ.ZS|||Banco Mundial/UNESCO|||%|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Média|||Proxy de qualidade de oferta educativa.
3|||Qualidade docente|||Rácio alunos/professor no secundário|||SE.SEC.ENRL.TC.ZS|||Banco Mundial/UNESCO|||alunos por professor|||-|||5|||80|||30|||Pesquisa externa/compilação local|||Média|||Proxy de pressão sobre a oferta educativa.
3|||Saúde e longevidade|||Esperança de vida saudável - HALE|||WHOSIS_000002|||OMS/GHO|||anos|||+|||30|||80|||60|||Pesquisa externa/compilação local|||Média|||Complementa esperança de vida com anos saudáveis.
4|||Proteção social|||Cobertura de proteção social e programas laborais|||per_allsp.cov_pop_tot|||Banco Mundial/ASPIRE|||% população|||+|||0|||100|||60|||Pesquisa externa/compilação local|||Alta|||Mede alcance efetivo da proteção social.
4|||Proteção social|||Cobertura de redes de segurança social no quintil mais pobre|||per_sa_allsa.cov_q1_tot|||Banco Mundial/ASPIRE|||% população|||+|||0|||100|||60|||Pesquisa externa/compilação local|||Alta|||Foco distributivo nos mais pobres.
4|||Proteção social|||Cobertura de seguro social no quintil mais pobre|||per_si_allsi.cov_q1_tot|||Banco Mundial/ASPIRE|||% população|||+|||0|||100|||40|||Pesquisa externa/compilação local|||Média|||Complementa assistência social.
4|||Pobreza multidimensional|||Pobreza multidimensional - World Bank headcount|||SI.POV.MPWB|||Banco Mundial/WDI|||% população|||-|||0|||100|||35|||Pesquisa externa/compilação local|||Alta|||Alternativa internacional ao IPM nacional.
4|||Pobreza multidimensional|||Pobreza multidimensional - UNDP headcount|||SI.POV.MPUN|||Banco Mundial/PNUD|||% população|||-|||0|||100|||35|||Pesquisa externa/compilação local|||Alta|||Benchmark PNUD para comparação internacional.
4|||Género|||Gender Development Index|||gdi|||PNUD/HDR|||0-1|||+|||0|||1|||0.95|||Pesquisa externa/compilação local|||Alta|||Indicador sintético de paridade no desenvolvimento humano.
4|||Género|||Gender Inequality Index|||gii|||PNUD/HDR|||0-1|||-|||0|||1|||0.35|||Pesquisa externa/compilação local|||Alta|||Indicador sintético de desigualdade de género.
4|||Desigualdade humana|||Coeficiente de desigualdade humana|||coef_ineq|||PNUD/HDR|||%|||-|||0|||100|||25|||Pesquisa externa/compilação local|||Média|||Complementa Gini/rendimento com educação e saúde.
4|||Inclusão financeira|||Conta em instituição financeira ou mobile money|||FX.OWN.TOTL.ZS|||Banco Mundial/Findex|||% adultos 15+|||+|||0|||100|||60|||Pesquisa externa/compilação local|||Alta|||Inclusão financeira é relevante para proteção e oportunidades.
4|||Inclusão financeira|||Conta financeira/mobile money - mulheres|||FX.OWN.TOTL.FE.ZS|||Banco Mundial/Findex|||% mulheres 15+|||+|||0|||100|||55|||Pesquisa externa/compilação local|||Alta|||Capta fosso de género no acesso financeiro.
4|||Inclusão financeira|||Conta financeira/mobile money - 40% mais pobres|||FX.OWN.TOTL.40.ZS|||Banco Mundial/Findex|||% adultos 15+|||+|||0|||100|||45|||Pesquisa externa/compilação local|||Alta|||Foco distributivo de inclusão financeira.
4|||Registo civil|||Registo de nascimento rural completo|||SP.REG.BRTH.RU.ZS|||Banco Mundial/UNICEF|||%|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Média|||Registo civil é pré-condição para acesso a direitos.
4|||Proteção e segurança das mulheres|||Violência doméstica contra mulheres|||-|||INE/IIMS 2023-2024|||%|||-|||0|||100|||10|||Ficheiro nacional/manual|||Alta|||Disponível em capítulos IIMS 2023-2024; útil para inclusão social/género.
5|||Energia|||Acesso rural à eletricidade|||EG.ELC.ACCS.RU.ZS|||Banco Mundial/IEA|||% população rural|||+|||0|||100|||50|||Pesquisa externa/compilação local|||Alta|||Permite medir desigualdade territorial no serviço.
5|||Energia|||Acesso urbano à eletricidade|||EG.ELC.ACCS.UR.ZS|||Banco Mundial/IEA|||% população urbana|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Alta|||Permite medir desigualdade territorial no serviço.
5|||Energia|||Eletricidade renovável no output elétrico|||EG.ELC.RNEW.ZS|||Banco Mundial/IEA|||% output elétrico|||+|||0|||100|||70|||Pesquisa externa/compilação local|||Média|||Relevante para matriz energética e resiliência.
5|||Energia|||Eletricidade hidroelétrica no output elétrico|||EG.ELC.HYRO.ZS|||Banco Mundial/IEA|||% output elétrico|||+/-|||0|||100|||-|||Pesquisa externa/compilação local|||Baixa|||Contexto de matriz energética; interpretar com resiliência a secas.
5|||Água|||Água potável básica rural|||SH.H2O.BASW.RU.ZS|||Banco Mundial/JMP|||% população rural|||+|||0|||100|||65|||Pesquisa externa/compilação local|||Alta|||Capta lacuna rural.
5|||Água|||Água potável básica urbana|||SH.H2O.BASW.UR.ZS|||Banco Mundial/JMP|||% população urbana|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Alta|||Capta qualidade da provisão urbana.
5|||Saneamento|||Saneamento básico rural|||SH.STA.BASS.RU.ZS|||Banco Mundial/JMP|||% população rural|||+|||0|||100|||45|||Pesquisa externa/compilação local|||Alta|||Capta lacuna rural.
5|||Saneamento|||Saneamento básico urbano|||SH.STA.BASS.UR.ZS|||Banco Mundial/JMP|||% população urbana|||+|||0|||100|||75|||Pesquisa externa/compilação local|||Alta|||Capta qualidade da provisão urbana.
5|||Cozinha limpa|||Energia limpa para cozinhar rural|||EG.CFT.ACCS.RU.ZS|||Banco Mundial/OMS|||% população rural|||+|||0|||100|||25|||Pesquisa externa/compilação local|||Alta|||Conecta infraestrutura, saúde e ambiente.
5|||Cozinha limpa|||Energia limpa para cozinhar urbana|||EG.CFT.ACCS.UR.ZS|||Banco Mundial/OMS|||% população urbana|||+|||0|||100|||60|||Pesquisa externa/compilação local|||Alta|||Conecta infraestrutura, saúde e ambiente.
5|||Transportes e logística|||Tráfego portuário de contentores|||IS.SHP.GOOD.TU|||Banco Mundial/WDI|||TEU|||+|||0|||-|||-|||Pesquisa externa/compilação local|||Média|||Proxy operacional de logística e comércio.
5|||Transportes e logística|||Passageiros transportados por via aérea|||IS.AIR.PSGR|||Banco Mundial/WDI|||número|||+|||0|||-|||-|||Pesquisa externa/compilação local|||Baixa|||Indicador de conectividade aérea.
5|||Transportes e logística|||Frete aéreo transportado|||IS.AIR.GOOD.MT.K1|||Banco Mundial/WDI|||milhões ton-km|||+|||0|||-|||-|||Pesquisa externa/compilação local|||Baixa|||Indicador de conectividade logística aérea.
5|||Investimento em infraestruturas|||Investimento privado em transportes|||IE.PPI.TRAN.CD|||Banco Mundial/PPI|||USD correntes|||+|||0|||-|||-|||Pesquisa externa/compilação local|||Baixa|||Pode ser volátil; usar como sinal de pipeline de investimento.
5|||Habitação e serviços|||Agregados com acesso a eletricidade/água/saneamento - Censo 2024|||-|||INE/RGPH 2024|||%|||+|||0|||100|||70|||Ficheiro nacional/manual|||Alta|||Disponível nos ficheiros censitários nacionais e provinciais no diretório local INE.
6|||Estrutura setorial|||Emprego na agricultura|||SL.AGR.EMPL.ZS|||Banco Mundial/OIT|||% emprego total|||-|||0|||100|||35|||Pesquisa externa/compilação local|||Média|||Proxy de transformação estrutural; interpretar com produtividade agrícola.
6|||Estrutura setorial|||Emprego na indústria|||SL.IND.EMPL.ZS|||Banco Mundial/OIT|||% emprego total|||+|||0|||100|||20|||Pesquisa externa/compilação local|||Média|||Capta absorção de emprego produtivo fora da agricultura.
6|||Estrutura setorial|||Emprego nos serviços|||SL.SRV.EMPL.ZS|||Banco Mundial/OIT|||% emprego total|||+|||0|||100|||45|||Pesquisa externa/compilação local|||Média|||Capta transição estrutural para serviços.
6|||Participação laboral|||Participação feminina na força de trabalho - OIT modelada|||SL.TLF.CACT.FE.ZS|||Banco Mundial/OIT|||%|||+|||0|||100|||65|||Pesquisa externa/compilação local|||Alta|||Complementa igualdade económica.
6|||Participação laboral|||Participação masculina na força de trabalho - OIT modelada|||SL.TLF.CACT.MA.ZS|||Banco Mundial/OIT|||%|||+|||0|||100|||75|||Pesquisa externa/compilação local|||Média|||Série de referência para comparação de género.
6|||Desemprego|||Desemprego nacional estimado|||SL.UEM.TOTL.NE.ZS|||Banco Mundial/OIT|||%|||-|||0|||60|||20|||Pesquisa externa/compilação local|||Alta|||Alternativa nacional ao desemprego modelado.
6|||Desemprego|||Desemprego feminino|||SL.UEM.TOTL.FE.ZS|||Banco Mundial/OIT|||%|||-|||0|||60|||20|||Pesquisa externa/compilação local|||Alta|||Capta assimetrias de género.
6|||Desemprego|||Desemprego masculino|||SL.UEM.TOTL.MA.ZS|||Banco Mundial/OIT|||%|||-|||0|||60|||18|||Pesquisa externa/compilação local|||Média|||Capta assimetrias de género.
6|||Juventude|||Desemprego juvenil feminino|||SL.UEM.1524.FE.ZS|||Banco Mundial/OIT|||%|||-|||0|||80|||30|||Pesquisa externa/compilação local|||Alta|||Juventude é uma dimensão crítica para Angola.
6|||Juventude|||Desemprego juvenil masculino|||SL.UEM.1524.MA.ZS|||Banco Mundial/OIT|||%|||-|||0|||80|||28|||Pesquisa externa/compilação local|||Média|||Juventude é uma dimensão crítica para Angola.
6|||Qualidade do emprego|||Emprego vulnerável feminino|||SL.EMP.VULN.FE.ZS|||Banco Mundial/OIT|||% emprego feminino|||-|||0|||100|||45|||Pesquisa externa/compilação local|||Alta|||Capta precariedade com lente de género.
6|||Qualidade do emprego|||Trabalhadoras assalariadas|||SL.EMP.WORK.FE.ZS|||Banco Mundial/OIT|||% emprego feminino|||+|||0|||100|||45|||Pesquisa externa/compilação local|||Média|||Aproxima formalidade e estabilidade de rendimento.
6|||Produtividade|||Produto por trabalhador - OIT|||GDP_205U_NOC_NB|||OIT/ILOSTAT|||USD constantes 2015|||+|||0|||-|||-|||Pesquisa externa/compilação local|||Alta|||Indicador de produtividade laboral.
6|||Formalização|||Emprego com contrato escrito / formal|||-|||INE/IEA|||%|||+|||0|||100|||50|||Ficheiro nacional/manual|||Alta|||Disponível nos ficheiros do Inquérito sobre Emprego em Angola.
7|||Segurança alimentar|||Insegurança alimentar severa|||SN.ITK.SVFI.ZS|||Banco Mundial/FAO|||% população|||-|||0|||100|||20|||Pesquisa externa/compilação local|||Alta|||Complementa insegurança moderada ou severa.
7|||Nutrição infantil|||Atraso de crescimento infantil|||SH.STA.STNT.ZS|||Banco Mundial/OMS/UNICEF|||% crianças <5|||-|||0|||60|||20|||Pesquisa externa/compilação local|||Alta|||Indicador central de nutrição e capital humano futuro.
7|||Nutrição infantil|||Emaciação infantil|||SH.STA.WAST.ZS|||Banco Mundial/OMS/UNICEF|||% crianças <5|||-|||0|||25|||5|||Pesquisa externa/compilação local|||Alta|||Indicador de desnutrição aguda.
7|||Despesa em saúde|||Despesa direta das famílias em saúde|||SH.XPD.OOPC.CH.ZS|||Banco Mundial/OMS|||% despesa corrente em saúde|||-|||0|||100|||20|||Pesquisa externa/compilação local|||Alta|||Capta proteção financeira em saúde.
7|||Despesa em saúde|||Despesa pública doméstica em saúde|||SH.XPD.GHED.GD.ZS|||Banco Mundial/OMS|||% PIB|||+|||0|||10|||3|||Pesquisa externa/compilação local|||Média|||Complementa despesa corrente total.
7|||Imunização|||Vacinação contra sarampo|||SH.IMM.MEAS|||Banco Mundial/OMS/UNICEF|||% crianças 12-23 meses|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Alta|||Cobertura essencial de saúde infantil.
7|||Imunização|||Vacinação HepB3|||SH.IMM.HEPB|||Banco Mundial/OMS/UNICEF|||% crianças 1 ano|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Média|||Cobertura essencial de saúde infantil.
7|||Doenças transmissíveis|||Incidência de malária|||SH.MLR.INCD.P3|||Banco Mundial/OMS|||por 1.000 em risco|||-|||0|||500|||100|||Pesquisa externa/compilação local|||Alta|||Doença crítica para Angola.
7|||Doenças transmissíveis|||Mortes estimadas por malária|||MALARIA_EST_DEATHS|||OMS/GHO|||número|||-|||0|||-|||0|||Pesquisa externa/compilação local|||Alta|||Doença crítica para Angola.
7|||Doenças transmissíveis|||Incidência de tuberculose|||SH.TBS.INCD|||Banco Mundial/OMS|||por 100.000 hab.|||-|||0|||500|||100|||Pesquisa externa/compilação local|||Média|||Carga de doença transmissível.
7|||Doenças transmissíveis|||Sucesso do tratamento da tuberculose|||SH.TBS.CURE.ZS|||Banco Mundial/OMS|||%|||+|||0|||100|||90|||Pesquisa externa/compilação local|||Média|||Capta capacidade do sistema de saúde.
7|||Doenças transmissíveis|||Incidência de VIH 15-49|||SH.HIV.INCD.ZS|||Banco Mundial/UNAIDS|||por 1.000 não infetados|||-|||0|||10|||1|||Pesquisa externa/compilação local|||Média|||Carga de doença transmissível.
7|||Doenças não transmissíveis|||Obesidade adulta|||NCD_BMI_30A|||OMS/GHO|||%|||-|||0|||50|||10|||Pesquisa externa/compilação local|||Baixa|||Risco emergente de doença crónica.
7|||Produção alimentar|||Índice de produção agrícola/culturas|||AG.PRD.CROP.XD|||Banco Mundial/FAO|||2014-2016=100|||+|||50|||200|||120|||Pesquisa externa/compilação local|||Média|||Complementa disponibilidade de alimentos.
7|||Produção alimentar|||Índice de produção pecuária|||AG.PRD.LVSK.XD|||Banco Mundial/FAO|||2014-2016=100|||+|||50|||200|||120|||Pesquisa externa/compilação local|||Média|||Complementa disponibilidade de alimentos.
7|||Saúde reprodutiva|||Pré-natal, parto assistido e cuidados maternos - IIMS|||-|||INE/IIMS 2023-2024|||%|||+|||0|||100|||80|||Ficheiro nacional/manual|||Alta|||Capítulos IIMS de cuidados maternos e saúde infantil.
8|||Capacidades produtivas|||Productive Capacities Index - geral|||UNCTAD|US.PCI|M6080|Category=0|||UNCTADstat|||índice|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Benchmark UNCTAD apropriado para diversificação e capacidade produtiva.
8|||Capacidades produtivas|||Productive Capacities Index - mudança estrutural|||UNCTAD|US.PCI|M6080|Category=10|||UNCTADstat|||índice|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Capta mudança estrutural.
8|||Capacidades produtivas|||Productive Capacities Index - setor privado|||UNCTAD|US.PCI|M6080|Category=80|||UNCTADstat|||índice|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Capta capacidade do setor privado.
8|||Capacidades produtivas|||Productive Capacities Index - transportes|||UNCTAD|US.PCI|M6080|Category=70|||UNCTADstat|||índice|||+|||0|||100|||45|||Nova dimensão / compilação local|||Média|||Capta infraestrutura logística produtiva.
8|||Dependência de recursos|||Rendas petrolíferas|||NY.GDP.PETR.RT.ZS|||Banco Mundial/WDI|||% PIB|||-|||0|||60|||15|||Nova dimensão / compilação local|||Alta|||Relevante para Angola por dependência petrolífera.
8|||Dependência de recursos|||Rendas totais de recursos naturais|||NY.GDP.TOTL.RT.ZS|||Banco Mundial/WDI|||% PIB|||-|||0|||70|||20|||Nova dimensão / compilação local|||Alta|||Capta dependência geral de recursos.
8|||Estrutura produtiva|||Indústria transformadora no PIB|||NV.IND.MANF.ZS|||Banco Mundial/WDI|||% PIB|||+|||0|||30|||12|||Nova dimensão / compilação local|||Alta|||Diversificação fora do petróleo.
8|||Estrutura produtiva|||Agricultura, florestas e pescas no PIB|||NV.AGR.TOTL.ZS|||Banco Mundial/WDI|||% PIB|||+/-|||0|||50|||-|||Nova dimensão / compilação local|||Média|||Contexto de transformação estrutural; positivo se associado a produtividade.
8|||Estrutura produtiva|||Serviços no PIB|||NV.SRV.TOTL.ZS|||Banco Mundial/WDI|||% PIB|||+|||0|||80|||45|||Nova dimensão / compilação local|||Média|||Diversificação setorial.
8|||Comércio externo|||Combustíveis nas exportações de mercadorias|||TX.VAL.FUEL.ZS.UN|||Banco Mundial/WDI|||% exportações mercadorias|||-|||0|||100|||70|||Nova dimensão / compilação local|||Alta|||Concentração petrolífera das exportações.
8|||Comércio externo|||Manufaturas nas exportações de mercadorias|||TX.VAL.MANF.ZS.UN|||Banco Mundial/WDI|||% exportações mercadorias|||+|||0|||100|||15|||Nova dimensão / compilação local|||Alta|||Diversificação exportadora.
8|||Comércio externo|||Exportações de alta tecnologia|||TX.VAL.TECH.MF.ZS|||Banco Mundial/WDI|||% exportações manufaturadas|||+|||0|||100|||5|||Nova dimensão / compilação local|||Média|||Sinal de sofisticação exportadora.
8|||Investimento e setor privado|||Formação bruta de capital fixo|||NE.GDI.FTOT.ZS|||Banco Mundial/WDI|||% PIB|||+|||0|||60|||25|||Nova dimensão / compilação local|||Alta|||Investimento produtivo.
8|||Investimento e setor privado|||IDE líquido total|||BX.KLT.DINV.WD.GD.ZS|||Banco Mundial/WDI|||% PIB|||+|||-10|||20|||5|||Nova dimensão / compilação local|||Alta|||Atração de investimento externo.
8|||Investimento e setor privado|||Crédito doméstico ao setor privado|||FS.AST.PRVT.GD.ZS|||Banco Mundial/WDI|||% PIB|||+|||0|||100|||35|||Nova dimensão / compilação local|||Alta|||Financiamento ao setor privado.
8|||Empreendedorismo|||Densidade de novas empresas|||IC.BUS.NDNS.ZS|||Banco Mundial/WDI|||novas empresas por 1.000 adultos 15-64|||+|||0|||20|||3|||Nova dimensão / compilação local|||Média|||Dinamismo empresarial formal.
8|||Diversificação nacional|||Crescimento do PIB não petrolífero|||-|||INE/MEP/BNA|||%|||+|||-10|||15|||5|||Ficheiro nacional/manual|||Alta|||Indicador central para Angola; há dados nas contas nacionais/relatórios BNA, mas requer extração harmonizada.
9|||Benchmark ambiental|||Environmental Performance Index - score|||-|||Yale/Columbia EPI|||0-100|||+|||0|||100|||55|||Nova dimensão / benchmark manual|||Média|||Benchmark internacional para saúde ambiental, ecossistemas e clima.
9|||Adaptação climática|||ND-GAIN Country Index - prontidão/resiliência|||-|||Notre Dame Global Adaptation Initiative|||índice|||+|||0|||100|||50|||Nova dimensão / benchmark manual|||Alta|||Relevante para vulnerabilidade climática e capacidade de resposta.
9|||Adaptação climática|||ND-GAIN - vulnerabilidade|||-|||Notre Dame Global Adaptation Initiative|||índice|||-|||0|||1|||0.35|||Nova dimensão / benchmark manual|||Alta|||Complementa prontidão climática.
9|||Emissões|||CO2 per capita excluindo LULUCF|||EN.GHG.CO2.PC.CE.AR5|||Banco Mundial/WDI|||t CO2e/capita|||-|||0|||20|||2|||Nova dimensão / compilação local|||Média|||Pegada de carbono per capita.
9|||Emissões|||Gases de efeito estufa per capita excluindo LULUCF|||EN.GHG.ALL.PC.CE.AR5|||Banco Mundial/WDI|||t CO2e/capita|||-|||0|||25|||4|||Nova dimensão / compilação local|||Média|||Pegada climática ampla.
9|||Emissões|||Emissões totais de GEE excluindo LULUCF|||EN.GHG.ALL.MT.CE.AR5|||Banco Mundial/WDI|||Mt CO2e|||-|||0|||-|||-|||Nova dimensão / compilação local|||Baixa|||Indicador de escala; usar com cautela em normalização.
9|||Qualidade ambiental|||Exposição média a PM2.5|||EN.ATM.PM25.MC.M3|||Banco Mundial/WDI|||microgramas/m3|||-|||0|||100|||15|||Nova dimensão / compilação local|||Alta|||Saúde ambiental e poluição.
9|||Energia limpa|||Consumo de energia renovável|||EG.FEC.RNEW.ZS|||Banco Mundial/WDI|||% consumo final|||+|||0|||100|||65|||Nova dimensão / compilação local|||Média|||Matriz energética e transição.
9|||Biodiversidade|||Áreas terrestres e marinhas protegidas|||ER.PTD.TOTL.ZS|||Banco Mundial/WDI|||% território|||+|||0|||30|||17|||Nova dimensão / compilação local|||Média|||Conservação e biodiversidade.
9|||Biodiversidade|||Área florestal|||AG.LND.FRST.ZS|||Banco Mundial/WDI|||% área terrestre|||+|||0|||80|||45|||Nova dimensão / compilação local|||Média|||Capital natural.
9|||Recursos hídricos|||Stress hídrico|||ER.H2O.FWST.ZS|||Banco Mundial/WDI|||% recursos hídricos|||-|||0|||100|||20|||Nova dimensão / compilação local|||Alta|||Relevante para clima, agricultura e território.
9|||Recursos hídricos|||Recursos internos renováveis de água per capita|||ER.H2O.INTR.PC|||Banco Mundial/WDI|||m3 per capita|||+|||0|||-|||-|||Nova dimensão / compilação local|||Baixa|||Indicador de dotação hídrica.
9|||Sustentabilidade económica|||Poupança líquida ajustada|||NY.ADJ.SVNG.GN.ZS|||Banco Mundial/WDI|||% RNB|||+|||-50|||50|||10|||Nova dimensão / compilação local|||Alta|||Integra capital natural e sustentabilidade económica.
9|||Sustentabilidade económica|||Depleção de recursos naturais|||NY.ADJ.DRES.GN.ZS|||Banco Mundial/WDI|||% RNB|||-|||0|||60|||10|||Nova dimensão / compilação local|||Alta|||Pressão sobre capital natural.
9|||Sustentabilidade económica|||Depleção de energia|||NY.ADJ.DNGY.GN.ZS|||Banco Mundial/WDI|||% RNB|||-|||0|||60|||8|||Nova dimensão / compilação local|||Média|||Importante para economia petrolífera.
9|||Riscos climáticos|||População afetada por secas, cheias e temperaturas extremas|||EN.CLC.MDAT.ZS|||Banco Mundial/WDI|||% população média 1990-2009|||-|||0|||100|||5|||Nova dimensão / compilação local|||Média|||Proxy de exposição climática histórica.
9|||Instituições ambientais|||CPIA - sustentabilidade ambiental|||IQ.CPA.ENVR.XQ|||Banco Mundial/WDI|||1-6|||+|||1|||6|||4|||Nova dimensão / compilação local|||Média|||Instituições e políticas ambientais.
10|||Estrutura demográfica|||População total|||SP.POP.TOTL|||Banco Mundial/WDI|||número|||+/-|||0|||-|||-|||Nova dimensão / compilação local|||Alta|||Indicador de escala; não deve ser normalizado como desempenho sem transformação.
10|||Estrutura demográfica|||Crescimento populacional|||SP.POP.GROW|||Banco Mundial/WDI|||% anual|||+/-|||0|||6|||-|||Nova dimensão / compilação local|||Alta|||Pressão de serviços e dinâmica demográfica.
10|||Urbanização|||População urbana|||SP.URB.TOTL.IN.ZS|||Banco Mundial/WDI|||% população|||+/-|||0|||100|||-|||Nova dimensão / compilação local|||Alta|||Estrutura territorial.
10|||Urbanização|||População rural|||SP.RUR.TOTL.ZS|||Banco Mundial/WDI|||% população|||+/-|||0|||100|||-|||Nova dimensão / compilação local|||Média|||Estrutura territorial.
10|||Densidade e ocupação|||Densidade populacional|||EN.POP.DNST|||Banco Mundial/WDI|||pessoas/km2|||+/-|||0|||-|||-|||Nova dimensão / compilação local|||Média|||Escala territorial e planeamento.
10|||Dependência demográfica|||Rácio de dependência total|||SP.POP.DPND|||Banco Mundial/WDI|||% população em idade ativa|||-|||0|||120|||60|||Nova dimensão / compilação local|||Alta|||Pressão sobre população ativa.
10|||Dependência demográfica|||Rácio de dependência jovem|||SP.POP.DPND.YG|||Banco Mundial/WDI|||% população em idade ativa|||-|||0|||120|||55|||Nova dimensão / compilação local|||Alta|||Pressão sobre educação, saúde e emprego futuro.
10|||Dependência demográfica|||Rácio de dependência idosa|||SP.POP.DPND.OL|||Banco Mundial/WDI|||% população em idade ativa|||+/-|||0|||50|||-|||Nova dimensão / compilação local|||Baixa|||Contexto de envelhecimento.
10|||Fecundidade|||Taxa de fecundidade total|||SP.DYN.TFRT.IN|||Banco Mundial/WDI|||nascimentos por mulher|||-|||1|||8|||3.5|||Nova dimensão / compilação local|||Alta|||Dinâmica demográfica e pressão sobre serviços.
10|||Fecundidade|||Fecundidade adolescente|||SP.ADO.TFRT|||Banco Mundial/WDI|||nascimentos por 1.000 mulheres 15-19|||-|||0|||250|||60|||Nova dimensão / compilação local|||Alta|||Cruza educação, saúde e género.
10|||Mobilidade|||Migração líquida|||SM.POP.NETM|||Banco Mundial/WDI|||número|||+/-|||-|||-|||-|||Nova dimensão / compilação local|||Média|||Contexto migratório.
10|||Mobilidade|||Refugiados por país de origem|||SM.POP.RHCR.EO|||Banco Mundial/UNHCR|||número|||-|||0|||-|||0|||Nova dimensão / compilação local|||Baixa|||Contexto de deslocação externa.
10|||Mobilidade|||Refugiados acolhidos por Angola|||SM.POP.RHCR.EA|||Banco Mundial/UNHCR|||número|||+/-|||0|||-|||-|||Nova dimensão / compilação local|||Baixa|||Contexto humanitário e regional.
10|||Habitação urbana|||População urbana em bairros precários/slums|||EN.POP.SLUM.UR.ZS|||Banco Mundial/UN-Habitat|||% população urbana|||-|||0|||100|||30|||Nova dimensão / compilação local|||Alta|||Urbanização e qualidade habitacional.
10|||Coesão territorial|||Dispersão do PIB per capita provincial|||-|||INE - PIB Anual por Província|||coeficiente/índice|||-|||0|||-|||-|||Ficheiro nacional/manual|||Alta|||Construível a partir do ficheiro local PIB Anual por Província.xlsx.
10|||Coesão territorial|||Dispersão da pobreza multidimensional municipal|||-|||INE - Pobreza Multidimensional nos Municípios|||coeficiente/índice|||-|||0|||-|||-|||Ficheiro nacional/manual|||Alta|||Construível a partir dos ficheiros locais de IPM municipal/provincial.
10|||Território e população|||Peso populacional de Luanda e principais polos urbanos|||-|||INE/RGPH 2024|||% população|||+/-|||0|||100|||-|||Ficheiro nacional/manual|||Média|||Indicador de concentração territorial; Censo 2024 disponível localmente.
11|||Benchmark digital|||Network Readiness Index - score|||-|||Network Readiness Index|||0-100|||+|||0|||100|||45|||Nova dimensão / benchmark manual|||Alta|||Benchmark de tecnologia, pessoas, governação e impacto digital.
11|||Governo digital|||UN E-Government Development Index - EGDI|||-|||UN DESA EGDI|||0-1|||+|||0|||1|||0.55|||Nova dimensão / benchmark manual|||Alta|||Benchmark de governo digital e serviços online.
11|||Governo digital|||GovTech Maturity Index|||GTMI|||Banco Mundial/GovTech|||0-1 / grupo|||+|||0|||1|||0.55|||Nova dimensão / benchmark manual|||Média|||Maturidade GovTech do setor público.
11|||Acesso digital|||Utilizadores de internet|||IT.NET.USER.ZS|||Banco Mundial/ITU|||% população|||+|||0|||100|||50|||Nova dimensão / compilação local|||Alta|||Indicador base de inclusão digital.
11|||Acesso digital|||Utilizadores de internet - mulheres|||IT.NET.USER.FE.ZS|||Banco Mundial/ITU|||% mulheres|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Capta fosso digital de género.
11|||Acesso digital|||Utilizadores de internet - homens|||IT.NET.USER.MA.ZS|||Banco Mundial/ITU|||% homens|||+|||0|||100|||55|||Nova dimensão / compilação local|||Média|||Capta fosso digital de género.
11|||Conectividade|||Subscrições móveis|||IT.CEL.SETS.P2|||Banco Mundial/ITU|||por 100 habitantes|||+|||0|||160|||100|||Nova dimensão / compilação local|||Alta|||Conectividade móvel.
11|||Conectividade|||Banda larga fixa|||IT.NET.BBND.P2|||Banco Mundial/ITU|||por 100 habitantes|||+|||0|||50|||8|||Nova dimensão / compilação local|||Alta|||Infraestrutura digital fixa.
11|||Segurança digital|||Servidores seguros de internet|||IT.NET.SECR.P6|||Banco Mundial/Netcraft|||por 1 milhão hab.|||+|||0|||-|||-|||Nova dimensão / compilação local|||Média|||Proxy de presença digital segura.
11|||Inclusão financeira digital|||Conta financeira ou mobile money|||FX.OWN.TOTL.ZS|||Banco Mundial/Findex|||% adultos 15+|||+|||0|||100|||60|||Nova dimensão / compilação local|||Alta|||Base para serviços financeiros digitais.
11|||Inclusão financeira digital|||Conta financeira/mobile money - jovens|||FX.OWN.TOTL.YG.ZS|||Banco Mundial/Findex|||% jovens 15-24|||+|||0|||100|||50|||Nova dimensão / compilação local|||Média|||Inclusão financeira da juventude.
11|||Inclusão financeira digital|||Conta financeira/mobile money - 40% mais pobres|||FX.OWN.TOTL.40.ZS|||Banco Mundial/Findex|||% adultos 15+|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Inclusão financeira digital distributiva.
11|||Capacidade digital|||Productive Capacities Index - ICT|||UNCTAD|US.PCI|M6080|Category=60|||UNCTADstat|||índice|||+|||0|||100|||45|||Nova dimensão / compilação local|||Alta|||Capacidade produtiva digital.
11|||Comércio digital|||Exportações de serviços TIC|||BX.GSR.CCIS.ZS|||Banco Mundial/WDI|||% exportações de serviços|||+|||0|||100|||10|||Nova dimensão / compilação local|||Média|||Sofisticação de serviços exportados.
11|||Comércio digital|||Exportações de bens TIC|||TX.VAL.ICTG.ZS.UN|||Banco Mundial/WDI|||% exportações de bens|||+|||0|||30|||3|||Nova dimensão / compilação local|||Baixa|||Sofisticação de bens exportados.
11|||Inovação|||Despesa em investigação e desenvolvimento|||GB.XPD.RSDV.GD.ZS|||Banco Mundial/UNESCO|||% PIB|||+|||0|||5|||1|||Nova dimensão / compilação local|||Média|||Input de inovação.
11|||Inovação|||Investigadores em I&D|||SP.POP.SCIE.RD.P6|||Banco Mundial/UNESCO|||por milhão de habitantes|||+|||0|||-|||-|||Nova dimensão / compilação local|||Média|||Capacidade científica.
11|||Inovação|||Pedidos de patente por residentes|||IP.PAT.RESD|||Banco Mundial/WIPO|||número|||+|||0|||-|||-|||Nova dimensão / compilação local|||Baixa|||Output formal de inovação.
11|||Inovação|||Pedidos de marca por residentes|||IP.TMK.RSCT|||Banco Mundial/WIPO|||número|||+|||0|||-|||-|||Nova dimensão / compilação local|||Baixa|||Dinamismo empresarial/branding.
11|||Pagamentos digitais|||Pagamentos eletrónicos e mobile payments - EMIS/BNA|||-|||BNA/EMIS|||número/valor/%|||+|||0|||-|||-|||Ficheiro nacional/manual|||Alta|||Indicador muito relevante para Angola; requer fonte nacional/EMIS ou relatórios BNA.
"""


BENCHMARKS = [
    ("Human Development Index / HDR", "PNUD", "Saúde, educação e rendimento; base para Capital Humano e Inclusão.", "https://hdr.undp.org/data-center"),
    ("Global SDG Indicator Framework", "ONU", "Cobertura ampla de indicadores de desenvolvimento sustentável e disponibilidade estatística.", "https://unstats.un.org/sdgs/indicators/indicators-list/"),
    ("Worldwide Governance Indicators", "Banco Mundial", "Seis dimensões de governação usadas no capítulo de governança.", "https://www.worldbank.org/en/publication/worldwide-governance-indicators"),
    ("Ibrahim Index of African Governance", "Mo Ibrahim Foundation", "Benchmark africano para segurança, estado de direito, participação, direitos, oportunidades e desenvolvimento humano.", "https://iiag.online/"),
    ("World Development Indicators / Data API", "Banco Mundial", "Fonte primária de séries anuais para Angola usadas em vários capítulos.", "https://data.worldbank.org/country/angola"),
    ("Statistical Performance Indicators", "Banco Mundial", "Qualidade/capacidade estatística, relevante para governança de dados.", "https://www.worldbank.org/en/programs/statistical-performance-indicators"),
    ("Social Progress Index", "Social Progress Imperative", "Benchmark de necessidades humanas básicas, fundamentos de bem-estar e oportunidade.", "https://www.socialprogress.org/social-progress-index"),
    ("Legatum Prosperity Index", "Legatum Institute", "Pilares amplos de prosperidade; usado como benchmark transversal.", "https://www.prosperity.com/about/resources"),
    ("Global Gender Gap Index", "World Economic Forum", "Benchmark de género em participação económica, educação, saúde e empowerment político.", "https://www.weforum.org/publications/global-gender-gap-report-2025/"),
    ("ND-GAIN Country Index", "Notre Dame Global Adaptation Initiative", "Vulnerabilidade climática e prontidão para adaptação.", "https://gain.nd.edu/our-work/country-index/"),
    ("Environmental Performance Index", "Yale/Columbia", "Saúde ambiental, vitalidade dos ecossistemas e política climática.", "https://epi.yale.edu/"),
    ("Network Readiness Index", "Portulans Institute", "Tecnologia, pessoas, governação e impacto digital.", "https://networkreadinessindex.org/"),
    ("UN E-Government Development Index", "UN DESA", "Governo digital e serviços públicos online.", "https://publicadministration.un.org/egovkb/en-us/Data-Center"),
    ("GovTech Maturity Index", "Banco Mundial", "Maturidade GovTech e digitalização do setor público.", "https://www.worldbank.org/en/programs/govtech/gtmi"),
    ("Global Health Observatory", "OMS", "Séries de saúde, mortalidade, cobertura e despesa em saúde.", "https://www.who.int/data/gho"),
    ("ILOSTAT", "OIT", "Emprego, desemprego, participação laboral e produtividade.", "https://ilostat.ilo.org/data/"),
    ("FAOSTAT", "FAO", "Produção alimentar e indicadores agrícolas.", "https://www.fao.org/faostat/"),
    ("UNCTADstat / Productive Capacities Index", "UNCTAD", "Diversificação e capacidades produtivas.", "https://unctadstat.unctad.org/"),
    ("INE Angola", "Instituto Nacional de Estatística", "Ficheiros locais de censo, IIMS, emprego, pobreza, contas nacionais e comércio.", "https://www.ine.gov.ao/"),
    ("BNA Estatísticas", "Banco Nacional de Angola", "Ficheiros locais de inflação, reservas, contas externas, crédito, juros e estatísticas monetárias.", "https://www.bna.ao/"),
]


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    return None if s in ("", "-") else s


def num(v):
    v = clean(v)
    if v is None:
        return None
    try:
        return float(v)
    except ValueError:
        return v


def cov(values, years):
    return sum(1 for y in years if values.get(y) not in (None, ""))


def covstr(values, years):
    return f"{cov(values, years)}/{len(years)}"


def last_year(values):
    for y in reversed(YEARS):
        if values.get(y) not in (None, ""):
            return int(y)
    return None


def infer_url(code, source, existing=None):
    code = clean(code)
    source = (clean(source) or "").lower()
    if code:
        if code.startswith("UNCTAD|"):
            return "https://unctadstat.unctad.org/"
        if code in {"hdi", "ihdi", "gdi", "gii", "eys", "mys", "coef_ineq", "ineq_edu"}:
            return "https://hdr.undp.org/data-center"
        if code.startswith(("WH", "MDG_", "MALARIA_", "GHED_", "NCD_", "HIV_", "SA_", "SDG")):
            return "https://www.who.int/data/gho"
        if code.startswith("FAO:"):
            return "https://www.fao.org/faostat/"
        if code in {"RL.EST", "GE.EST", "PV.EST", "RQ.EST", "CC.EST", "VA.EST"}:
            return "https://www.worldbank.org/en/publication/worldwide-governance-indicators"
        if code.startswith(("GDP_", "EAP_", "EMP_", "UNE_")):
            return "https://ilostat.ilo.org/data/"
        if code.startswith("AFDB|"):
            return "https://dataportal.opendataforafrica.org/"
        if code == "GTMI":
            return "https://www.worldbank.org/en/programs/govtech/gtmi"
        if re.fullmatch(r"[A-Za-z0-9_.]+", code) or code.startswith("per_"):
            return f"https://data.worldbank.org/indicator/{code}?locations=AO"
    if "ine" in source:
        return "https://www.ine.gov.ao/"
    if "bna" in source or "emis" in source:
        return "https://www.bna.ao/"
    if "transparency" in source:
        return "https://www.transparency.org/en/cpi"
    if "ibrahim" in source:
        return "https://iiag.online/"
    if "budget" in source:
        return "https://internationalbudget.org/open-budget-survey/"
    if "enterprise" in source:
        return "https://www.enterprisesurveys.org/"
    if "nd-gain" in source or "notre dame" in source:
        return "https://gain.nd.edu/our-work/country-index/"
    if "epi" in source or "environmental performance" in source:
        return "https://epi.yale.edu/"
    if "network readiness" in source:
        return "https://networkreadinessindex.org/"
    if "egdi" in source or "e-government" in source:
        return "https://publicadministration.un.org/egovkb/en-us/Data-Center"
    if "govtech" in source:
        return "https://www.worldbank.org/en/programs/govtech/gtmi"
    return existing


def custom_code(indicator, source=None):
    name = (indicator or "").lower()
    source = (source or "").lower()
    if "percepção da corrupção" in name or "percepcao da corrupcao" in name:
        return "TI_CPI_SCORE"
    if "índice de desenvolvimento humano" in name or "indice de desenvolvimento humano" in name:
        return "hdi"
    if "anos médios de escolaridade" in name or "anos medios de escolaridade" in name:
        return "mys"
    if "anos esperados de escolaridade" in name:
        return "eys"
    if "pobreza multidimensional" in name and "incid" in name:
        return "SI.POV.MPUN"
    if "transparency" in source and "corrup" in name:
        return "TI_CPI_SCORE"
    return None


def load_v4_ids():
    ids = {}
    if not V4_FILE.exists():
        return ids
    wb = load_workbook(V4_FILE, read_only=True, data_only=True)
    ws = wb["Indicadores_v4"]
    headers = [c.value for c in ws[4]]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row and row[0]:
            d = dict(zip(headers, row))
            ids[(clean(d.get("Dimensão")), clean(d.get("Indicador")))] = clean(d.get("ID"))
    return ids


def load_lookup():
    lookup, meta = {}, {}
    wb = load_workbook(DATA_FILE, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if sheet in ("Índice", "Metodologia"):
            continue
        ws = wb[sheet]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        if "Indicator Code" not in headers:
            continue
        ci = headers.index("Indicator Code")
        ni = headers.index("Indicator Name") if "Indicator Name" in headers else None
        si = headers.index("Source") if "Source" in headers else None
        ui = headers.index("Unit") if "Unit" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = clean(row[ci] if row else None)
            if not code or code in lookup:
                continue
            values = {y: row[headers.index(y)] if y in headers else None for y in YEARS}
            lookup[code] = values
            meta[code] = {
                "sheet": sheet,
                "name": clean(row[ni]) if ni is not None else None,
                "source": clean(row[si]) if si is not None else sheet,
                "unit": clean(row[ui]) if ui is not None else None,
            }
    return lookup, meta


def status(values, code):
    if cov(values, YEARS):
        if cov(values, [str(y) for y in range(2000, 2015)]):
            return "Com série automática na compilação local"
        return "Com série 2015-2025 na base IGDA/local"
    if clean(code):
        return "Código/API identificado; sem série local carregada"
    return "Recolha manual / fonte externa ou ficheiro nacional"


def norm_key(dim, indicator, code):
    code = clean(code)
    if code:
        return (dim, "CODE", code)
    return (dim, "NAME", re.sub(r"\s+", " ", (indicator or "").lower().strip()))


def add_row(rows, seen, seq, lookup, meta, v4_ids, *, dim, subtema, indicador, tipo, prioridade,
            sentido=None, unidade=None, minv=None, maxv=None, target=None, peso=1, code=None,
            fonte=None, note=None, origin=None, prefer_base=None):
    code = clean(code) or custom_code(indicador, fonte)
    key = norm_key(dim, indicador, code)
    if key in seen:
        return
    values = {y: None for y in YEARS}
    src_origin = origin or "Pesquisa externa"
    if code and code in lookup:
        values.update(lookup[code])
        src_origin = f"{DATA_FILE.name}:{meta[code]['sheet']}"
        fonte = fonte or meta[code].get("source")
        unidade = unidade or meta[code].get("unit")
        local_name = meta[code].get("name")
        if local_name and local_name.lower() != indicador.lower():
            note = ((note + " | ") if note else "") + f"Nome local/API: {local_name}"
    if prefer_base:
        for y in BASE_YEARS:
            if prefer_base.get(y) not in (None, ""):
                values[y] = prefer_base[y]
    seq[dim] += 1
    rid = v4_ids.get((dim, indicador)) or f"{DIM_PREFIX[dim]}{seq[dim]:03d}"
    row = {
        "ID": rid,
        "Dimensão": dim,
        "Subtema": subtema,
        "Indicador": indicador,
        "Tipo": tipo,
        "Prioridade": prioridade,
        "Sentido": sentido,
        "Unidade": unidade,
        "Mínimo": minv,
        "Máximo": maxv,
        "Meta 2027": target,
        "Peso sugerido": peso,
        "Código/API": code,
        "Fonte": fonte,
        "Fonte_URL": infer_url(code, fonte),
        "Estado da disponibilidade": status(values, code),
        "Cobertura 2000-2025": covstr(values, YEARS),
        "Cobertura 2015-2025": covstr(values, BASE_YEARS),
        "Último ano obs.": last_year(values),
        "Origem no ficheiro": src_origin,
        "Justificação / nota": note,
        **values,
    }
    rows.append(row)
    seen.add(key)


def build_rows():
    rows, seen, seq = [], set(), defaultdict(int)
    v4_ids = load_v4_ids()
    lookup, meta = load_lookup()

    wb = load_workbook(BASE_FILE, read_only=True, data_only=True)
    ws = wb["Base_Alargada"]
    headers = [c.value for c in ws[4]]
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not raw or raw[0] is None:
            continue
        b = dict(zip(headers, raw))
        dim = clean(b.get("Dimensão"))
        if dim not in DIM_SHEETS:
            continue
        code = clean(b.get("Código")) or custom_code(clean(b.get("Indicador")), clean(b.get("Fonte")))
        origin = f"{BASE_FILE.name}:Base_Alargada"
        if code in meta:
            origin += f" + {DATA_FILE.name}:{meta[code]['sheet']}"
        add_row(
            rows, seen, seq, lookup, meta, v4_ids,
            dim=dim,
            subtema=clean(b.get("Subíndice curto")),
            indicador=clean(b.get("Indicador")),
            tipo=clean(b.get("Tipo")) or "Atual/sugerido no IGDA",
            prioridade=clean(b.get("Prioridade")) or "Alta",
            sentido=clean(b.get("Sentido")),
            unidade=clean(b.get("Unidade")),
            minv=b.get("Mínimo"),
            maxv=b.get("Máximo"),
            target=b.get("Meta 2027"),
            peso=b.get("Peso") or 1,
            code=code,
            fonte=clean(b.get("Fonte")),
            note=clean(b.get("Racional / nota")),
            origin=origin,
            prefer_base=b,
        )

    for line in EXTRAS.strip().splitlines():
        parts = [clean(p) for p in line.split("|||")]
        if len(parts) != 13:
            raise ValueError(f"Bad EXTRAS line with {len(parts)} fields: {line}")
        dnum, subtema, indicador, code, fonte, unidade, sentido, minv, maxv, target, tipo, prioridade, note = parts
        add_row(
            rows, seen, seq, lookup, meta, v4_ids,
            dim=DIM_BY_NUM[dnum],
            subtema=subtema,
            indicador=indicador,
            tipo=tipo,
            prioridade=prioridade,
            sentido=sentido,
            unidade=unidade,
            minv=num(minv),
            maxv=num(maxv),
            target=num(target),
            code=code,
            fonte=fonte,
            note=note,
        )

    prio_rank = {"Alta": 0, "Média": 1, "Media": 1, "Baixa": 2}
    dim_rank = {dim: i for i, dim in enumerate(DIM_SHEETS)}
    rows.sort(key=lambda r: (dim_rank.get(r["Dimensão"], 99), prio_rank.get(r.get("Prioridade"), 3), r.get("Subtema") or "", r.get("ID") or ""))
    return rows


def make_xlsx(rows):
    by_dim = defaultdict(list)
    for row in rows:
        by_dim[row["Dimensão"]].append(row)

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    newdim_fill = PatternFill("solid", fgColor="E2F0D9")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(*(Side(style="thin", color="D9E2F3"),) * 4)

    ws = wb.create_sheet("00_Indice")
    ws["A1"] = "IGDA-BDA - zoo de indicadores por dimensão"
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ws["A2"] = "Compilação de indicadores atuais, sugeridos, disponíveis na base local e candidatos de pesquisa externa para Angola."
    ws["A3"] = "Data de geração: 2026-06-02. Base de seleção; não implica ponderação final nem cálculo automático."
    idx_headers = ["Dimensão", "Folha", "Nº indicadores", "Com dados 2000-2025", "Com dados 2015-2025", "Manuais/candidatos", "Alta prioridade"]
    for j, h in enumerate(idx_headers, 1):
        c = ws.cell(5, j, h)
        c.fill = header_fill
        c.font = white_font
    rr = 6
    for dim, sheet in DIM_SHEETS.items():
        dr = by_dim[dim]
        vals = [
            dim, sheet, len(dr),
            sum(1 for r in dr if cov({y: r.get(y) for y in YEARS}, YEARS)),
            sum(1 for r in dr if cov({y: r.get(y) for y in BASE_YEARS}, BASE_YEARS)),
            sum(1 for r in dr if "manual" in (r.get("Estado da disponibilidade") or "").lower() or "sem série" in (r.get("Estado da disponibilidade") or "").lower()),
            sum(1 for r in dr if r.get("Prioridade") == "Alta"),
        ]
        for j, v in enumerate(vals, 1):
            ws.cell(rr, j, v)
        rr += 1
    ws.cell(rr + 2, 1, "Fontes locais usadas").font = Font(bold=True)
    ws.cell(rr + 3, 1, str(BASE_FILE))
    ws.cell(rr + 4, 1, str(V4_FILE))
    ws.cell(rr + 5, 1, str(DATA_FILE))
    for col, width in {"A": 45, "B": 24, "C": 18, "D": 22, "E": 22, "F": 22, "G": 18}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"

    def write_sheet(name, data, title=None):
        ws = wb.create_sheet(name)
        start = 1
        if title:
            ws.cell(1, 1, title).font = Font(bold=True, size=13, color="1F4E78")
            start = 3
        for j, h in enumerate(HEADERS, 1):
            c = ws.cell(start, j, h)
            c.fill = header_fill
            c.font = white_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for i, row in enumerate(data, start + 1):
            is_new = str(row["Dimensão"]).startswith(("8.", "9.", "10.", "11."))
            is_manual = "manual" in (row.get("Estado da disponibilidade") or "").lower() or "sem série" in (row.get("Estado da disponibilidade") or "").lower()
            for j, h in enumerate(HEADERS, 1):
                c = ws.cell(i, j, row.get(h))
                c.border = border
                if j <= 21:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    if is_manual:
                        c.fill = manual_fill
                    elif is_new:
                        c.fill = newdim_fill
                if h == "Fonte_URL" and isinstance(row.get(h), str) and row[h].startswith("http"):
                    c.hyperlink = row[h]
                    c.style = "Hyperlink"
                if j >= 22:
                    c.number_format = "0.00"
        ws.freeze_panes = ws.cell(start + 1, 4).coordinate
        ws.auto_filter.ref = f"A{start}:{get_column_letter(len(HEADERS))}{start + len(data)}"
        widths = {"A": 10, "B": 34, "C": 24, "D": 44, "E": 28, "F": 12, "G": 10, "H": 22, "M": 24, "N": 24, "O": 38, "P": 36, "T": 34, "U": 55}
        for idx in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(get_column_letter(idx), 12)

    write_sheet("Zoo_Total", rows, "Todos os indicadores candidatos")
    for dim, sheet in DIM_SHEETS.items():
        write_sheet(sheet, by_dim[dim], dim)

    ws = wb.create_sheet("Fontes_Benchmarks")
    bench_headers = ["Benchmark/Fonte", "Instituição", "Uso no zoo", "URL"]
    for j, h in enumerate(bench_headers, 1):
        c = ws.cell(1, j, h)
        c.fill = header_fill
        c.font = white_font
    for i, row in enumerate(BENCHMARKS, 2):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if j == 4:
                c.hyperlink = v
                c.style = "Hyperlink"
    for col, width in {"A": 34, "B": 24, "C": 70, "D": 55}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = "70AD47" if ws.title.startswith(("08", "09", "10", "11")) else "5B9BD5"
    wb.save(OUT_XLSX)


def make_xls(rows):
    by_dim = defaultdict(list)
    for row in rows:
        by_dim[row["Dimensão"]].append(row)
    book = xlwt.Workbook(encoding="utf-8")
    header = xlwt.easyxf("font: bold on, colour white; pattern: pattern solid, fore_colour dark_blue; align: horiz center, vert center, wrap on; borders: left thin, right thin, top thin, bottom thin")
    text = xlwt.easyxf("align: wrap on, vert top; borders: left thin, right thin, top thin, bottom thin")
    manual = xlwt.easyxf("pattern: pattern solid, fore_colour light_yellow; align: wrap on, vert top; borders: left thin, right thin, top thin, bottom thin")
    new = xlwt.easyxf("pattern: pattern solid, fore_colour light_green; align: wrap on, vert top; borders: left thin, right thin, top thin, bottom thin")
    title = xlwt.easyxf("font: bold on, height 260, colour dark_blue")
    numstyle = xlwt.easyxf("borders: left thin, right thin, top thin, bottom thin", num_format_str="0.00")

    ws = book.add_sheet("00_Indice", cell_overwrite_ok=True)
    ws.write(0, 0, "IGDA-BDA - zoo de indicadores por dimensão", title)
    ws.write(1, 0, "Compilação de indicadores atuais, sugeridos, disponíveis na base local e candidatos de pesquisa externa para Angola.", text)
    idx_headers = ["Dimensão", "Folha", "Nº indicadores", "Com dados 2000-2025", "Com dados 2015-2025", "Manuais/candidatos", "Alta prioridade"]
    for j, h in enumerate(idx_headers):
        ws.write(4, j, h, header)
    rr = 5
    for dim, sheet in DIM_SHEETS.items():
        dr = by_dim[dim]
        vals = [
            dim, sheet, len(dr),
            sum(1 for r in dr if cov({y: r.get(y) for y in YEARS}, YEARS)),
            sum(1 for r in dr if cov({y: r.get(y) for y in BASE_YEARS}, BASE_YEARS)),
            sum(1 for r in dr if "manual" in (r.get("Estado da disponibilidade") or "").lower() or "sem série" in (r.get("Estado da disponibilidade") or "").lower()),
            sum(1 for r in dr if r.get("Prioridade") == "Alta"),
        ]
        for j, v in enumerate(vals):
            ws.write(rr, j, v, text)
        rr += 1
    for j, width in enumerate([11000, 5000, 4000, 5200, 5200, 5200, 4200]):
        ws.col(j).width = width

    def write_sheet(name, data, sheet_title=None):
        ws = book.add_sheet(name[:31], cell_overwrite_ok=True)
        start = 0
        if sheet_title:
            ws.write(0, 0, sheet_title[:32760], title)
            start = 2
        for j, h in enumerate(HEADERS):
            ws.write(start, j, h, header)
        for i, row in enumerate(data, start + 1):
            is_new = str(row["Dimensão"]).startswith(("8.", "9.", "10.", "11."))
            is_manual = "manual" in (row.get("Estado da disponibilidade") or "").lower() or "sem série" in (row.get("Estado da disponibilidade") or "").lower()
            st = manual if is_manual else (new if is_new else text)
            for j, h in enumerate(HEADERS):
                v = row.get(h)
                if v is None:
                    v = ""
                if isinstance(v, str):
                    v = v[:32760]
                ws.write(i, j, v, numstyle if j >= 21 and isinstance(v, (int, float)) else st)
        for j in range(len(HEADERS)):
            ws.col(j).width = 8500 if j in [1, 3, 14, 15, 19] else (12000 if j == 20 else (4200 if j < 21 else 2800))
        ws.set_panes_frozen(True)
        ws.set_horz_split_pos(start + 1)
        ws.set_vert_split_pos(3)

    write_sheet("Zoo_Total", rows, "Todos os indicadores candidatos")
    for dim, sheet in DIM_SHEETS.items():
        write_sheet(sheet, by_dim[dim], dim)

    ws = book.add_sheet("Fontes_Benchmarks", cell_overwrite_ok=True)
    for j, h in enumerate(["Benchmark/Fonte", "Instituição", "Uso no zoo", "URL"]):
        ws.write(0, j, h, header)
    for i, row in enumerate(BENCHMARKS, 1):
        for j, v in enumerate(row):
            ws.write(i, j, v, text)
    for j, width in enumerate([9000, 6500, 16000, 14000]):
        ws.col(j).width = width
    book.save(str(OUT_XLS))


def main():
    rows = build_rows()
    make_xlsx(rows)
    make_xls(rows)
    by_dim = defaultdict(list)
    for row in rows:
        by_dim[row["Dimensão"]].append(row)
    print(f"Gerado: {OUT_XLS.resolve()}")
    print(f"Gerado: {OUT_XLSX.resolve()}")
    print(f"Total indicadores: {len(rows)}")
    for dim, sheet in DIM_SHEETS.items():
        dr = by_dim[dim]
        data_count = sum(1 for r in dr if cov({y: r.get(y) for y in YEARS}, YEARS))
        manual_count = sum(1 for r in dr if "manual" in (r.get("Estado da disponibilidade") or "").lower() or "sem série" in (r.get("Estado da disponibilidade") or "").lower())
        print(f"{sheet}: {len(dr)} indicadores | com dados: {data_count} | manuais/sem serie: {manual_count}")


if __name__ == "__main__":
    main()
